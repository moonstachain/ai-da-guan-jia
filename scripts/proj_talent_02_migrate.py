#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REPO_ROOT = Path("/Users/liming/Documents/codex-ai-gua-jia-01")
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from dashboard.feishu_deploy import FeishuBitableAPI
from scripts.create_kangbo_signal_tables import load_feishu_credentials

WIKI_TOKEN = "LO0uweNbDiVl6QkTUmgc3SfWnmf"
BASE_NAME = "原力OS_高潜猎聘管理"
APP_TOKEN = "UeWubgnADaLw3tsjIB8cPup8ndf"
TRACKER_APP_TOKEN = "PVDgbdWYFaDLBiss0hlcM5WRnQc"
TRACKER_TABLE_ID = "tblB9JQ4cROTBUnr"

T01_TABLE_NAME = "T01_候选人主表"
T02_TABLE_NAME = "T02_AI简历分析"
T04_TABLE_NAME = "T04_面试记录"
T08_TABLE_NAME = "T08_操作日志"

LOCAL_TZ = timezone(timedelta(hours=8))
DEFAULT_SOURCE = Path("/Users/liming/Downloads/【专家大脑】_印经小院-人才库 (1).xlsx")
DATE_FORMAT = "%Y-%m-%d"


def now_ms() -> int:
    return int(datetime.now(LOCAL_TZ).timestamp() * 1000)


def now_date() -> str:
    return datetime.now(LOCAL_TZ).strftime(DATE_FORMAT)


def run_id() -> str:
    return f"adagj-proj-talent-02-{datetime.now(LOCAL_TZ).strftime('%Y%m%d-%H%M%S')}"


def jwrite(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def twrite(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text.rstrip() + "\n", encoding="utf-8")


def req(api: FeishuBitableAPI, method: str, path: str, payload: dict[str, Any] | None = None) -> dict[str, Any]:
    return api._request(path, method=method, payload=payload)


def table_by_name(api: FeishuBitableAPI, app_token: str, table_name: str) -> dict[str, Any]:
    resp = req(api, "GET", f"/open-apis/bitable/v1/apps/{app_token}/tables?page_size=500")
    for item in (resp.get("data") or {}).get("items") or []:
        if str(item.get("name") or item.get("table_name") or "").strip() == table_name:
            return item
    raise RuntimeError(f"missing table: {table_name}")


def records(api: FeishuBitableAPI, app_token: str, table_id: str) -> list[dict[str, Any]]:
    resp = req(api, "GET", f"/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/records?page_size=500")
    return (resp.get("data") or {}).get("items") or []


def create_record(api: FeishuBitableAPI, app_token: str, table_id: str, fields: dict[str, Any]) -> str:
    resp = req(api, "POST", f"/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/records", {"fields": fields})
    record = (resp.get("data") or {}).get("record") or {}
    record_id = str(record.get("record_id") or record.get("id") or "").strip()
    if not record_id:
        items = (resp.get("data") or {}).get("records") or []
        if items:
            record_id = str(items[0].get("record_id") or items[0].get("id") or "").strip()
    if not record_id:
        raise RuntimeError(f"failed to create record in {table_id}")
    return record_id


def batch_create_records(api: FeishuBitableAPI, app_token: str, table_id: str, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    req(api, "POST", f"/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/records/batch_create", {"records": [{"fields": row} for row in rows]})


def upsert_tracker_row(api: FeishuBitableAPI, payload: dict[str, Any]) -> str:
    resp = req(api, "GET", f"/open-apis/bitable/v1/apps/{TRACKER_APP_TOKEN}/tables/{TRACKER_TABLE_ID}/records?page_size=500")
    items = (resp.get("data") or {}).get("items") or []
    existing = next((row for row in items if str((row.get("fields") or {}).get("task_id") or "").strip() == payload["task_id"]), None)
    if existing:
        record_id = str(existing.get("record_id") or existing.get("id") or "").strip()
        req(api, "PUT", f"/open-apis/bitable/v1/apps/{TRACKER_APP_TOKEN}/tables/{TRACKER_TABLE_ID}/records/{record_id}", {"fields": payload})
        return "updated"
    req(api, "POST", f"/open-apis/bitable/v1/apps/{TRACKER_APP_TOKEN}/tables/{TRACKER_TABLE_ID}/records", {"fields": payload})
    return "created"


def extract_first_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    match = re.search(r"(\d+)", str(value))
    return int(match.group(1)) if match else None


def extract_grade_letter(value: Any) -> str | None:
    if value in (None, ""):
        return None
    match = re.search(r"([SABCD])级", str(value))
    if match:
        return match.group(1)
    match = re.search(r"\b([SABCD])\b", str(value))
    return match.group(1) if match else None


def map_human_decision(value: Any) -> str:
    text = str(value or "").strip()
    if text == "面试":
        return "通过"
    if text == "拒绝":
        return "拒绝"
    if text == "观察":
        return "待定"
    return "待定"


def map_stage(human_decision: str, transcript: str | None) -> str:
    if transcript and transcript.strip():
        return "S3一面"
    if human_decision == "通过":
        return "S3一面"
    if human_decision == "拒绝":
        return "已淘汰"
    return "S1简历筛选"


def merge_text(*parts: Any) -> str:
    chunks = [str(part).strip() for part in parts if str(part or "").strip()]
    return "\n\n---\n\n".join(chunks)


def non_url_attachment(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if "http://" in text or "https://" in text:
        return text
    return ""


def parse_rows(workbook_path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(workbook_path, read_only=False, data_only=True)
    ws = wb.active
    rows: list[dict[str, Any]] = []
    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row_idx, 1).value
        if name in (None, ""):
            continue
        rows.append(
            {
                "source_row": row_idx,
                "name": str(name).strip(),
                "resume_text": ws.cell(row_idx, 2).value,
                "resume_screenshot": ws.cell(row_idx, 3).value,
                "prompt_output": ws.cell(row_idx, 4).value,
                "ai_resume_analysis": ws.cell(row_idx, 5).value,
                "prompt_thinking": ws.cell(row_idx, 6).value,
                "score_raw": ws.cell(row_idx, 7).value,
                "grade_raw": ws.cell(row_idx, 8).value,
                "human_decision_raw": ws.cell(row_idx, 9).value,
                "recommended_questions": ws.cell(row_idx, 10).value,
                "recommended_questions_ai": ws.cell(row_idx, 11).value,
                "transcript": ws.cell(row_idx, 12).value,
                "interview_ai_thinking": ws.cell(row_idx, 13).value,
                "interview_ai_analysis": ws.cell(row_idx, 14).value,
                "stage_raw": ws.cell(row_idx, 15).value,
            }
        )
    return rows


def build_t01_fields(row: dict[str, Any], migration_ts: int) -> dict[str, Any]:
    human_decision = map_human_decision(row["human_decision_raw"])
    stage = map_stage(human_decision, str(row["transcript"] or ""))
    fields = {
        "姓名": row["name"],
        "目标项目": "八万四千",
        "目标岗位": "私域运营负责人",
        "来源渠道": "Boss直聘",
        "简历解析文本": str(row["resume_text"] or "").strip(),
        "AI综合评分": extract_first_int(row["score_raw"]) or "",
        "AI等级": extract_grade_letter(row["grade_raw"]) or "",
        "AI推荐意见": str(row["grade_raw"] or "").strip(),
        "人类决策": human_decision,
        "当前阶段": stage,
        "阶段更新时间": migration_ts,
        "创建时间": migration_ts,
        "标签": ["社招"],
    }
    screenshot = non_url_attachment(row["resume_screenshot"])
    if screenshot:
        fields["简历截图"] = screenshot
    return {key: value for key, value in fields.items() if value not in ("", None, [], {})}


def build_t02_fields(row: dict[str, Any], migration_ts: int, candidate_record_id: str) -> dict[str, Any]:
    fields = {
        "候选人": [candidate_record_id],
        "使用提示词版本": "v0-legacy-印经小院",
        "AI思考过程": str(row["prompt_thinking"] or "").strip(),
        "AI输出结果": merge_text(row["prompt_output"], row["ai_resume_analysis"]),
        "推荐面试问题": merge_text(row["recommended_questions"], row["recommended_questions_ai"]),
        "分析时间": migration_ts,
    }
    return {key: value for key, value in fields.items() if value not in ("", None, [], {})}


def build_t04_fields(row: dict[str, Any], migration_ts: int, candidate_record_id: str) -> dict[str, Any]:
    transcript = str(row["transcript"] or "").strip()
    if not transcript:
        return {}
    fields = {
        "候选人": [candidate_record_id],
        "面试轮次": "一面",
        "面试日期": migration_ts,
        "逐字稿文本": transcript,
        "AI面试分析": str(row["interview_ai_analysis"] or "").strip(),
    }
    interview_thinking = str(row["interview_ai_thinking"] or "").strip()
    if interview_thinking and not fields["AI面试分析"]:
        fields["AI面试分析"] = interview_thinking
    return {key: value for key, value in fields.items() if value not in ("", None, [], {})}


def field_count_report(rows: list[dict[str, Any]]) -> dict[str, Any]:
    counts = Counter()
    counts["source_rows"] = len(rows)
    counts["transcripts"] = sum(1 for row in rows if str(row.get("transcript") or "").strip())
    counts["missing_screenshots"] = sum(1 for row in rows if not non_url_attachment(row.get("resume_screenshot")))
    counts["decision_approve"] = sum(1 for row in rows if map_human_decision(row.get("human_decision_raw")) == "通过")
    counts["decision_reject"] = sum(1 for row in rows if map_human_decision(row.get("human_decision_raw")) == "拒绝")
    counts["decision_pending"] = sum(1 for row in rows if map_human_decision(row.get("human_decision_raw")) == "待定")
    counts["stage_s3"] = sum(1 for row in rows if map_stage(map_human_decision(row.get("human_decision_raw")), str(row.get("transcript") or "")) == "S3一面")
    counts["stage_rejected"] = sum(1 for row in rows if map_stage(map_human_decision(row.get("human_decision_raw")), str(row.get("transcript") or "")) == "已淘汰")
    counts["stage_s1"] = sum(1 for row in rows if map_stage(map_human_decision(row.get("human_decision_raw")), str(row.get("transcript") or "")) == "S1简历筛选")
    return dict(counts)


def ensure_empty_or_force(api: FeishuBitableAPI, app_token: str, table_id: str, *, force: bool) -> None:
    existing = records(api, app_token, table_id)
    if existing and not force:
        raise RuntimeError(f"{table_id} already has {len(existing)} records; rerun with --force if you intentionally want to append.")


def main() -> int:
    parser = argparse.ArgumentParser(description="Migrate PROJ-TALENT-02 candidate data from xlsx into Feishu.")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--force", action="store_true", help="Allow writes even when target tables already have data.")
    args = parser.parse_args()

    source_path: Path = args.source
    if not source_path.exists():
        raise FileNotFoundError(source_path)

    rows = parse_rows(source_path)
    if not rows:
        raise RuntimeError("No candidate rows found in source workbook.")

    migration_stamp = now_ms()
    migration_day = now_date()
    rid = run_id()
    run_dir = REPO_ROOT / "artifacts" / "ai-da-guan-jia" / "runs" / migration_day / rid
    run_dir.mkdir(parents=True, exist_ok=True)

    creds = load_feishu_credentials()
    api = FeishuBitableAPI(creds["app_id"], creds["app_secret"])
    node = req(api, "GET", f"/open-apis/wiki/v2/spaces/get_node?token={WIKI_TOKEN}")
    app_token = str(((node.get("data") or {}).get("node") or {}).get("obj_token") or "").strip()
    if not app_token:
        raise RuntimeError("Unable to resolve app token from wiki token.")
    if app_token != APP_TOKEN:
        print(f"warning: resolved app token {app_token} differs from expected {APP_TOKEN}")

    base = table_by_name(api, app_token, T01_TABLE_NAME)
    if str(base.get("table_id") or "").strip() != "tblGLmRbQ5E8kXGD":
        print(f"warning: unexpected T01 table id {base.get('table_id')}")

    t01 = table_by_name(api, app_token, T01_TABLE_NAME)
    t02 = table_by_name(api, app_token, T02_TABLE_NAME)
    t04 = table_by_name(api, app_token, T04_TABLE_NAME)
    t08 = table_by_name(api, app_token, T08_TABLE_NAME)
    t01_id = str(t01.get("table_id") or "").strip()
    t02_id = str(t02.get("table_id") or "").strip()
    t04_id = str(t04.get("table_id") or "").strip()
    t08_id = str(t08.get("table_id") or "").strip()

    for table_id in [t01_id, t02_id, t04_id]:
        ensure_empty_or_force(api, app_token, table_id, force=args.force)

    source_to_t01: dict[int, str] = {}
    t01_rows: list[dict[str, Any]] = []
    creation_notes: list[dict[str, Any]] = []
    for row in rows:
        fields = build_t01_fields(row, migration_stamp)
        record_id = create_record(api, app_token, t01_id, fields)
        source_to_t01[int(row["source_row"])] = record_id
        t01_rows.append({"source_row": row["source_row"], "record_id": record_id, "name": row["name"], "fields": fields})
        creation_notes.append({"source_row": row["source_row"], "record_id": record_id, "name": row["name"]})

    t02_rows: list[dict[str, Any]] = []
    for row in rows:
        record_id = source_to_t01[int(row["source_row"])]
        fields = build_t02_fields(row, migration_stamp, record_id)
        t02_record_id = create_record(api, app_token, t02_id, fields)
        t02_rows.append({"source_row": row["source_row"], "record_id": t02_record_id, "candidate_record_id": record_id, "fields": fields})

    t04_rows: list[dict[str, Any]] = []
    for row in rows:
        fields = build_t04_fields(row, migration_stamp, source_to_t01[int(row["source_row"])])
        if not fields:
            continue
        t04_record_id = create_record(api, app_token, t04_id, fields)
        t04_rows.append({"source_row": row["source_row"], "record_id": t04_record_id, "candidate_record_id": source_to_t01[int(row["source_row"])], "fields": fields})

    t08_payload = {
        "操作时间": migration_stamp,
        "操作人": "Codex",
        "操作类型": "数据迁移",
        "目标表": "PROJ-TALENT-02",
        "目标记录": "batch_migration_20",
        "操作详情": f"从 {source_path.name} 迁移 {len(rows)} 条候选人记录到 T01/T02/T04；T04 实际写入 {len(t04_rows)} 条。",
    }
    t08_record_id = create_record(api, app_token, t08_id, t08_payload)

    tracker_payload = {
        "task_id": "PROJ-TALENT-02",
        "project_id": "PROJ-TALENT",
        "project_name": "高潜力候选人猎聘管理",
        "project_status": "进行中",
        "task_name": "现有数据迁移",
        "task_status": "已完成",
        "priority": "P0",
        "owner": "Codex",
        "start_date": migration_stamp,
        "completion_date": migration_stamp,
        "blockers": "",
        "evidence_ref": str(run_dir),
        "dependencies": "PROJ-TALENT-01",
        "notes": f"已迁移 {len(rows)} 条候选人记录，T04 写入 {len(t04_rows)} 条，T08 新增操作日志 1 条。",
    }
    tracker_action = upsert_tracker_row(api, tracker_payload)

    report = {
        "run_id": rid,
        "run_dir": str(run_dir),
        "source_path": str(source_path),
        "source_sheet": "印经小院-人才库",
        "source_rows": len(rows),
        "migration_timestamp_ms": migration_stamp,
        "migration_day": migration_day,
        "app_token": app_token,
        "tables": {
            "T01_候选人主表": t01_id,
            "T02_AI简历分析": t02_id,
            "T04_面试记录": t04_id,
            "T08_操作日志": t08_id,
        },
        "counts": {
            "T01_created": len(t01_rows),
            "T02_created": len(t02_rows),
            "T04_created": len(t04_rows),
            "T08_created": 1,
            "tracker_action": tracker_action,
        },
        "source_statistics": field_count_report(rows),
        "source_to_t01": creation_notes,
        "t01_rows": t01_rows,
        "t02_rows": t02_rows,
        "t04_rows": t04_rows,
        "t08_record_id": t08_record_id,
        "self_evaluation": {
            "gained": "把 xlsx 里的 20 条候选人按源行顺序写入 Feishu，并把 T02/T04 的关联链路打通。",
            "wasted": "最初误读了 workbook 结构，差点把空白表当成 blocker。",
            "iterate_next": "如果后续要重跑，最好先加一层幂等标识或清理流程，避免重复写入。",
        },
    }

    report_path = run_dir / "migration_report.json"
    md_path = run_dir / "PROJ-TALENT-02-SUCCESS.md"
    jwrite(report_path, report)
    twrite(
        md_path,
        "\n".join(
            [
                "# PROJ-TALENT-02 Success",
                "",
                f"- Source: `{source_path}`",
                f"- Rows migrated: `{len(rows)}`",
                f"- T04 rows created: `{len(t04_rows)}`",
                f"- T08 record: `{t08_record_id}`",
                f"- Tracker action: `{tracker_action}`",
                "",
                "## Notes",
                "",
                "- Screenshot values in the workbook were non-URL attachment placeholders, so they were intentionally skipped.",
                "- Candidate order follows the workbook row order, including duplicate names.",
            ]
        ),
    )
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
