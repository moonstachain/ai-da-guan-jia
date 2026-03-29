from __future__ import annotations

import argparse
import json
import os
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from dashboard.feishu_deploy import FeishuBitableAPI


WORKBOOK_PATH = Path("/Users/liming/Downloads/原力创业人生财富地图-2.xlsx")
REGISTRY_PATH = REPO_ROOT / "work/boomerang-miaoda-app-v1/server/wealth_map_table_registry.json"
WIKI_LINK = "https://h52xu4gwob.feishu.cn/wiki/ToyewusRYinRsbkburicq4ygn2c"


def row_values(ws, row_idx: int) -> list[Any]:
    row = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
    return [value for value in row if value is not None]


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def chunked(items: list[dict[str, Any]], size: int = 50) -> list[list[dict[str, Any]]]:
    return [items[i : i + size] for i in range(0, len(items), size)]


def load_registry() -> dict[str, Any]:
    return json.loads(REGISTRY_PATH.read_text(encoding="utf-8"))


def table_info(registry: dict[str, Any], table_key: str) -> dict[str, Any]:
    table = registry["tables"][table_key]
    return {
        "table_id": table["table_id"],
        "fields": table["fields"],
    }


def delete_all_records(api: FeishuBitableAPI, app_token: str, table_id: str) -> int:
    deleted = 0
    for record in api.list_records(app_token, table_id):
      record_id = str(record.get("record_id") or "").strip()
      if not record_id:
          continue
      api._request(f"/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/records/{record_id}", method="DELETE")
      deleted += 1
    return deleted


def seed_table(api: FeishuBitableAPI, app_token: str, table_id: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
    existing = delete_all_records(api, app_token, table_id)
    for batch in chunked(rows, 50):
        if batch:
            api.batch_create_records(app_token, table_id, batch)
    verified = len(api.list_records(app_token, table_id))
    return {"deleted": existing, "created": len(rows), "verified": verified}


def build_seed_data(wb):
    k1 = wb["K1_原力觉醒"]
    k2 = wb["K2_品类独创"]
    k3 = wb["K3_模式升维"]
    k4 = wb["K4_壁垒锁定"]
    l1 = wb["L1_十大标杆学员"]
    l2 = wb["L2_学员原子拆解"]

    t1_rows = []
    for row_idx in range(6, 11):
        dim, en, q, genius, great, quote, k1_case, k5_case = row_values(k1, row_idx)
        t1_rows.append(
            {
                "维度": text(dim),
                "英文名": text(en),
                "核心问题": text(q),
                "天才区信号": text(genius),
                "卓越区信号": text(great),
                "经典判词": text(quote),
                "康波K1案例": text(k1_case),
                "康波K5案例": text(k5_case),
            }
        )

    t2_rows = []
    for row_idx in range(6, 10):
        motto, en, logic, standard, student, history, anti, tips = row_values(k2, row_idx)
        t2_rows.append(
            {
                "真言": text(motto),
                "英文": text(en),
                "核心逻辑": text(logic),
                "判断标准": text(standard),
                "学员案例": text(student),
                "康波历史对标": text(history),
                "反面案例": text(anti),
                "操作要点": text(tips),
            }
        )

    t3_rows = []
    for row_idx in range(13, 19):
        direction, definition, market, persona, account, student, history, opp = row_values(k2, row_idx)
        t3_rows.append(
            {
                "人群方向": text(direction),
                "定义": text(definition),
                "市场特征": text(market),
                "典型用户画像": text(persona),
                "心理账户偏好": text(account),
                "学员适配案例": text(student),
                "康波对标案例": text(history),
                "关键机会点": text(opp),
            }
        )

    t4_rows = []
    for row_idx in range(6, 16):
        linkage, module, question, atom_count, student, history, ai, trap = row_values(k3, row_idx)
        t4_rows.append(
            {
                "链路": text(linkage),
                "九宫格模块": text(module),
                "核心问题": text(question),
                "原子组件数": text(atom_count),
                "学员典型打法": text(student),
                "康波历史经典案例": text(history),
                "AI时代升维方向": text(ai),
                "关键陷阱": text(trap),
            }
        )

    t5_rows = []
    for row_idx in range(6, 22):
        dim, idx, name, logic, standard, student, history, ai = row_values(k4, row_idx)
        t5_rows.append(
            {
                "维度": text(dim),
                "控制点序号": text(idx),
                "控制点名称": text(name),
                "核心逻辑": text(logic),
                "判断标准": text(standard),
                "学员案例": text(student),
                "康波历史经典": text(history),
                "AI时代新玩法": text(ai),
            }
        )

    t6_rows = []
    for idx, row_idx in enumerate(range(3, 12), start=1):
        case_row = row_values(l1, row_idx)
        (
            record_id,
            company,
            founder,
            archetype,
            industry,
            stage,
            revenue,
            awakening_adv,
            awakening_obs,
            awakening_shadow,
            category_motto,
            category_sweet,
            category_account,
            model_front,
            model_back,
            model_value,
            barrier_built,
            barrier_todo_ai,
        ) = case_row
        t6_rows.append(
            {
                "编号": text(record_id),
                "学员企业": text(company),
                "创始人": text(founder),
                "原型人格": text(archetype),
                "行业赛道": text(industry),
                "当前阶段": text(stage),
                "年营收规模": text(revenue),
                "原力觉醒_非对称优势": text(awakening_adv),
                "原力觉醒_非理性痴迷": text(awakening_obs),
                "原力觉醒_阴影转化": text(awakening_shadow),
                "品类独创_16字真言达成度": text(category_motto),
                "品类独创_甜用户": text(category_sweet),
                "品类独创_心理账户": text(category_account),
                "模式升维_前链路": text(model_front),
                "模式升维_后链路": text(model_back),
                "模式升维_财链路": text(model_value),
                "壁垒锁定_已建壁垒": text(barrier_built),
                "壁垒锁定_待建壁垒_AI介入": text(barrier_todo_ai),
            }
        )

    t7_rows = []
    for row_idx in range(22, 26):
        account, mindset, elasticity, ladder, pitch, student, history, path = row_values(k2, row_idx)
        t7_rows.append(
            {
                "心理账户": text(account),
                "用户心态": text(mindset),
                "定价弹性": text(elasticity),
                "价值层级": text(ladder),
                "经典话术": text(pitch),
                "学员案例": text(student),
                "康波对标": text(history),
                "升级路径": text(path),
            }
        )

    t8_rows = []
    for row_idx in range(29, 41):
        name, drive, genius, shadow, power, student, history, tracks = row_values(k1, row_idx)
        t8_rows.append(
            {
                "原型": text(name),
                "核心驱力": text(drive),
                "天才区": text(genius),
                "阴影面": text(shadow),
                "商业超能力": text(power),
                "代表学员": text(student),
                "康波历史代表": text(history),
                "适配赛道": text(tracks),
            }
        )

    return {
        "WM_T1_原力觉醒_非对称优势": t1_rows,
        "WM_T2_品类独创_16字真言": t2_rows,
        "WM_T3_品类独创_人群细分": t3_rows,
        "WM_T4_模式升维_三链路九宫格": t4_rows,
        "WM_T5_壁垒锁定_16控制点": t5_rows,
        "WM_T6_学员案例_全景诊断": t6_rows,
        "WM_T7_心理账户_价值阶梯": t7_rows,
        "WM_T8_12原型人格": t8_rows,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Seed the existing wealth-map Feishu tables from workbook 2.")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if not WORKBOOK_PATH.exists():
        raise SystemExit(f"Workbook not found: {WORKBOOK_PATH}")
    if not REGISTRY_PATH.exists():
        raise SystemExit(f"Registry not found: {REGISTRY_PATH}")

    registry = load_registry()
    app_token = str(registry["base"]["app_token"]).strip()
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    seed_data = build_seed_data(wb)

    api = FeishuBitableAPI(os.environ["FEISHU_APP_ID"], os.environ["FEISHU_APP_SECRET"])
    app_token_live = api.resolve_app_token(WIKI_LINK)
    if app_token_live != app_token:
        print(f"registry app_token {app_token} differs from wiki-resolved token {app_token_live}; using live token", flush=True)
        app_token = app_token_live

    summary: dict[str, Any] = {"app_token": app_token, "tables": {}}

    for table_name, rows in seed_data.items():
        table = registry["tables"][table_name]
        table_id = str(table["table_id"]).strip()
        if args.dry_run:
            summary["tables"][table_name] = {"table_id": table_id, "planned_records": len(rows)}
            continue
        result = seed_table(api, app_token, table_id, rows)
        summary["tables"][table_name] = {"table_id": table_id, **result}

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
