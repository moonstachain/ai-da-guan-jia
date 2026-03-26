#!/usr/bin/env python3
"""PROJ-V5-BASE-01: build the Yuanli OS Feishu base from the V2 xlsx spec."""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from dashboard.feishu_deploy import FeishuBitableAPI

try:
    from scripts.create_kangbo_signal_tables import load_feishu_credentials
except ModuleNotFoundError:  # pragma: no cover - direct script execution fallback
    from create_kangbo_signal_tables import load_feishu_credentials


TASK_ID = "PROJ-V5-BASE-01"
TARGET_BASE_NAME = "原力OS_一体化_V2"
WORKBOOK_PATH = Path("/Users/liming/Downloads/原力OS_飞书多维表_建表规格_V2.xlsx")
DEFAULT_LINK = "https://h52xu4gwob.feishu.cn/wiki/DJIVwbDv9iOJN5kJlRMcZiaynNe?from=from_copylink"
DEFAULT_ACCOUNT_ID = "feishu-claw"
SHANGHAI_TZ = timezone(timedelta(hours=8))

ARTIFACT_ROOT = REPO_ROOT / "artifacts" / "ai-da-guan-jia" / "runs"
FEISHU_ARTIFACT_DIR = REPO_ROOT / "artifacts" / "ai-da-guan-jia" / "feishu"
REGISTRY_PATH = FEISHU_ARTIFACT_DIR / "table_registry_v2.json"
PARTIAL_REGISTRY_PATH = FEISHU_ARTIFACT_DIR / "table_registry_v2_partial.json"
SCHEMA_MANIFEST_PATH = FEISHU_ARTIFACT_DIR / "table_schema_manifest_v2.json"
VIEW_CONTRACT_PATH = FEISHU_ARTIFACT_DIR / "table_views_v2.json"

FIELD_TYPE_MAP = {
    "单行文本": 1,
    "多行文本": 1,
    "数字": 2,
    "单选": 3,
    "多选": 4,
    "日期": 5,
    "复选框": 7,
    "URL": 15,
}

NUMBER_FORMAT_DEFAULT = "0"
NUMBER_FORMAT_DECIMAL = "0.0"

RETRYABLE_ERROR_CODES = {
    1254036,
    1254290,
    1254291,
    1254607,
    1255001,
    1255002,
    1255003,
    1255004,
    1255040,
}

DEFAULT_VIEW_NAME_BY_TABLE = {
    "T01_runtime_control": "总控视图",
    "T02_weekly_focus": "总控视图",
    "T03_strategic_tasks": "全部任务",
    "T04_tenants": "全部租户",
    "T05_role_templates": "全部模板",
    "T06_clone_instances": "全部实例",
    "T07_clone_run_logs": "全部运行",
    "T08_training_cycles": "全部轮次",
    "T09_capability_proposals": "全部提案",
    "T10_clone_scores": "全部评分",
    "T11_clone_risks": "全部风险",
    "T12_evolution_tracks": "全部轨迹",
    "T13_collaboration_logs": "全部协同",
    "T14_memory_health": "总控视图",
    "T15_governance_maturity": "审计历史",
    "T16_evolution_chronicle": "全部编年",
    "T17_decisions": "全部决策",
    "T18_skill_inventory": "全部Skill",
    "T19_component_heatmap": "全部热图",
}

P0_TABLE_KEYS = {
    "T01_runtime_control",
    "T02_weekly_focus",
    "T03_strategic_tasks",
    "T04_tenants",
    "T05_role_templates",
    "T06_clone_instances",
    "T14_memory_health",
}

KNOWN_OPTION_COLORS = {
    # common english states
    "healthy": 0,
    "active": 0,
    "completed": 0,
    "approved": 0,
    "resolved": 0,
    "accepted": 0,
    "up": 0,
    "stable": 1,
    "in_progress": 1,
    "planning": 1,
    "planned": 1,
    "draft": 1,
    "proposed": 1,
    "provisioning": 1,
    "consultation": 1,
    "training": 1,
    "review": 1,
    "degraded": 2,
    "blocked": 2,
    "medium": 2,
    "suspended": 2,
    "partial": 2,
    "warning": 2,
    "external": 2,
    "high": 3,
    "critical": 3,
    "failed": 3,
    "rejected": 3,
    "down": 3,
    "archived": 5,
    "deprecated": 5,
    "readonly": 5,
    "open": 3,
    # personas / orgs
    "claude": 1,
    "codex": 4,
    "human": 0,
    "employee": 0,
    "partner": 4,
    # service tiers / visibility
    "internal_core": 1,
    "internal_standard": 0,
    "client_premium": 4,
    "client_basic": 5,
    "hq_internal_full": 1,
    "hq_internal_limited": 2,
    "client_scoped": 4,
    # vectors and domains
    "control": 1,
    "task": 0,
    "clone": 2,
    "evidence": 4,
    "governance": 3,
    "v1_多机": 4,
    "v2_clone": 2,
    "v3_客户": 3,
    "内部": 1,
    # planning buckets
    "p0": 3,
    "p1": 2,
    "p2": 1,
    "p3": 5,
    "critical": 3,
    "important": 2,
    "nice_to_have": 5,
    # source and quadrant
    "github": 4,
    "local": 0,
    "core": 0,
    "有用": 1,
    "待观察": 2,
    "核心": 0,
    "已合并": 4,
    "淘汰": 5,
    "治理": 1,
    "执行": 0,
    "投研": 4,
    "运营": 2,
    "工具": 5,
    "其他": 1,
    # memory health
    "健康": 0,
    "亚健康": 2,
    "异常": 3,
    "离线": 5,
    # governance trend
    "flat": 1,
}

DECIMAL_NUMBER_FIELDS = {
    "avg_quality_score",
    "before_score",
    "after_score",
    "delta",
    "delivery_score",
    "accuracy_score",
    "efficiency_score",
    "collaboration_score",
    "overall_score",
}

DATE_ONLY_FIELDS = {
    "date",
    "audit_date",
    "start_date",
    "completion_date",
    "started_at",
    "completed_at",
    "resolved_at",
}

AUTO_FILL_DATE_FIELDS = {
    "updated_at",
    "generated_at",
    "created_at",
    "run_at",
    "scored_at",
}

VIEW_TYPE_MAP = {
    "表格": "grid",
    "看板": "kanban",
}

VIEW_FILTER_FIELD_ALIASES = {
    "status": "task_status",
}


@dataclass(frozen=True)
class FieldSpec:
    name: str
    source_type: str
    description: str
    default: Any
    required: bool
    options_text: str | None
    api_type: int
    options: tuple[str, ...] = ()


@dataclass(frozen=True)
class ViewSpec:
    table_key: str
    view_name: str
    view_type_cn: str
    filter_text: str
    group_sort_text: str
    service_scene: str

    @property
    def view_type(self) -> str:
        return VIEW_TYPE_MAP.get(self.view_type_cn.strip(), "grid")


@dataclass(frozen=True)
class TableSpec:
    table_key: str
    display_name: str
    domain: str
    purpose: str
    expected_rows: str
    plugin_key: str
    phase: str
    field_specs: tuple[FieldSpec, ...]
    default_view_name: str
    extra_views: tuple[ViewSpec, ...] = ()

    @property
    def field_count(self) -> int:
        return len(self.field_specs)


@dataclass
class TableLiveState:
    table_key: str
    table_id: str
    live_name: str
    field_ids_by_name: dict[str, str]
    field_meta_by_name: dict[str, dict[str, Any]]
    views_by_name: dict[str, dict[str, Any]]
    created: bool = False
    renamed: bool = False
    default_view_renamed: bool = False
    created_fields: list[str] = field(default_factory=list)
    updated_fields: list[str] = field(default_factory=list)
    created_views: list[str] = field(default_factory=list)
    patched_views: list[str] = field(default_factory=list)


def now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d-%H%M%S")


def shanghai_now() -> datetime:
    return datetime.now(tz=SHANGHAI_TZ)


def json_write(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def parse_error_code(exc: Exception) -> int | None:
    text = str(exc)
    if "Feishu API error" not in text:
        return None
    try:
        raw = text.split(": ", 1)[1]
        payload = json.loads(raw)
        code = payload.get("code")
        return int(code) if code is not None else None
    except Exception:
        return None


def call_feishu(
    api: FeishuBitableAPI,
    path: str,
    *,
    method: str = "GET",
    payload: dict[str, Any] | None = None,
    retries: int = 5,
    pause_seconds: float = 0.12,
) -> dict[str, Any]:
    last_exc: Exception | None = None
    for attempt in range(retries):
        try:
            response = api._request(path, method=method, payload=payload)
            if method.upper() in {"POST", "PUT", "PATCH", "DELETE"} and pause_seconds > 0:
                time.sleep(pause_seconds)
            return response
        except Exception as exc:  # pragma: no cover - live API only
            last_exc = exc
            code = parse_error_code(exc)
            if code in RETRYABLE_ERROR_CODES and attempt < retries - 1:
                time.sleep(pause_seconds * (attempt + 1))
                continue
            raise
    assert last_exc is not None
    raise last_exc


def split_options(text: str | None) -> tuple[str, ...]:
    if not text:
        return ()
    normalized = text.strip()
    if not normalized or normalized in {"无", "None", "none"}:
        return ()
    tokens = re.split(r"\s*/\s*|\s*、\s*|\s*,\s*|\s*;\s*|\s*\|\s*", normalized)
    return tuple(token.strip() for token in tokens if token.strip())


def normalize_name_token(value: str) -> str:
    return value.strip().lower()


def option_color(option_name: str, index: int) -> int:
    normalized = normalize_name_token(option_name)
    if normalized in KNOWN_OPTION_COLORS:
        return KNOWN_OPTION_COLORS[normalized]
    return [0, 1, 2, 3, 4, 5][index % 6]


def field_type_from_cn(field_type_cn: str) -> int:
    cleaned = field_type_cn.strip()
    if cleaned not in FIELD_TYPE_MAP:
        raise ValueError(f"Unsupported field type in workbook: {field_type_cn}")
    return FIELD_TYPE_MAP[cleaned]


def field_type_kind(field_type_cn: str) -> str:
    if field_type_cn in {"单行文本", "多行文本"}:
        return "text"
    if field_type_cn == "数字":
        return "number"
    if field_type_cn == "单选":
        return "single_select"
    if field_type_cn == "多选":
        return "multi_select"
    if field_type_cn == "日期":
        return "date"
    if field_type_cn == "复选框":
        return "checkbox"
    if field_type_cn == "URL":
        return "url"
    raise ValueError(f"Unsupported field type: {field_type_cn}")


def number_formatter(field_name: str) -> str:
    if field_name in DECIMAL_NUMBER_FIELDS:
        return NUMBER_FORMAT_DECIMAL
    return NUMBER_FORMAT_DEFAULT


def date_formatter(field_name: str) -> str:
    if field_name in DATE_ONLY_FIELDS or field_name.endswith("_date") or field_name == "date":
        return "yyyy/MM/dd"
    return "yyyy/MM/dd HH:mm"


def auto_fill_date(field_name: str) -> bool:
    return field_name in AUTO_FILL_DATE_FIELDS or field_name.endswith("_at")


def _normalized_row_values(sheet, row_idx: int) -> list[str]:
    values: list[str] = []
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(row_idx, col).value
        if value in (None, ""):
            continue
        values.append(str(value).strip())
    return values


def _detect_header_row(sheet) -> int:
    header_markers = (
        ("字段名 (field_name)", "字段类型"),
        ("表名", "视图名", "视图类型"),
    )
    max_probe_rows = min(sheet.max_row, 8)
    for row_idx in range(1, max_probe_rows + 1):
        values = set(_normalized_row_values(sheet, row_idx))
        for markers in header_markers:
            if all(marker in values for marker in markers):
                return row_idx
    return 2


def load_sheet_rows(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    header_row = _detect_header_row(sheet)
    headers = [sheet.cell(header_row, col).value for col in range(1, sheet.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        values = [sheet.cell(row_idx, col).value for col in range(1, sheet.max_column + 1)]
        if not any(value not in (None, "") for value in values):
            continue
        row: dict[str, Any] = {}
        for index, header in enumerate(headers):
            if header in (None, ""):
                continue
            row[str(header).strip()] = values[index]
        rows.append(row)
    return rows


def _row_value(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return None


def parse_field_specs(workbook_path: Path) -> list[TableSpec]:
    workbook = load_workbook(workbook_path, data_only=True)
    overview = workbook["00_建表总览"]
    view_rows = load_sheet_rows(workbook_path, "视图规划")
    views_by_table: dict[str, list[ViewSpec]] = {}
    for row in view_rows:
        table_key = str(row.get("表名") or "").strip()
        view_name = str(row.get("视图名") or "").strip()
        if not table_key or not view_name:
            continue
        views_by_table.setdefault(table_key, []).append(
            ViewSpec(
                table_key=table_key,
                view_name=view_name,
                view_type_cn=str(row.get("视图类型") or "").strip(),
                filter_text=str(row.get("过滤条件") or "").strip(),
                group_sort_text=str(row.get("分组/排序") or "").strip(),
                service_scene=str(row.get("服务场景") or "").strip(),
            )
        )

    table_meta_rows: list[dict[str, Any]] = []
    for row_idx in range(2, overview.max_row + 1):
        row = [overview.cell(row_idx, col).value for col in range(1, overview.max_column + 1)]
        if not any(value not in (None, "") for value in row):
            continue
        table_meta_rows.append(
            {
                "序号": row[0],
                "域": row[1],
                "表名": row[2],
                "表显示名": row[3],
                "用途说明": row[4],
                "预估初始行数": row[5],
                "核心capability插件": row[6],
            }
        )

    table_specs: list[TableSpec] = []
    for index, meta in enumerate(table_meta_rows, start=1):
        table_key = str(meta["表名"]).strip()
        sheet_rows = load_sheet_rows(workbook_path, table_key)
        field_specs: list[FieldSpec] = []
        for row in sheet_rows:
            field_name = str(_row_value(row, "field_name", "字段名 (field_name)") or "").strip()
            field_type_cn = str(_row_value(row, "字段类型", "field_type") or "").strip()
            if not field_name or not field_type_cn:
                continue
            options_text = _row_value(row, "选项/约束", "options")
            options = split_options(str(options_text)) if field_type_cn in {"单选", "多选"} else ()
            field_specs.append(
                FieldSpec(
                    name=field_name,
                    source_type=field_type_cn,
                    description=str(_row_value(row, "字段说明", "description") or "").strip(),
                    default=_row_value(row, "默认值", "default"),
                    required=str(_row_value(row, "必填", "required") or "").strip() == "是",
                    options_text=str(options_text).strip() if options_text not in (None, "") else None,
                    api_type=field_type_from_cn(field_type_cn),
                    options=options,
                )
            )

        if table_key not in DEFAULT_VIEW_NAME_BY_TABLE:
            raise ValueError(f"Missing default view mapping for {table_key}")

        table_specs.append(
            TableSpec(
                table_key=table_key,
                display_name=str(meta["表显示名"]).strip(),
                domain=str(meta["域"]).strip(),
                purpose=str(meta["用途说明"]).strip(),
                expected_rows=str(meta["预估初始行数"]).strip(),
                plugin_key=str(meta["核心capability插件"]).strip(),
                phase="P0" if table_key in P0_TABLE_KEYS else "P1",
                field_specs=tuple(field_specs),
                default_view_name=DEFAULT_VIEW_NAME_BY_TABLE[table_key],
                extra_views=tuple(
                    view
                    for view in views_by_table.get(table_key, [])
                    if view.view_name and view.view_name != DEFAULT_VIEW_NAME_BY_TABLE[table_key]
                ),
            )
        )

    return table_specs


def build_schema_manifest(table_specs: list[TableSpec]) -> dict[str, Any]:
    return {
        "task_id": TASK_ID,
        "base_name": TARGET_BASE_NAME,
        "workbook_path": str(WORKBOOK_PATH),
        "created_at": shanghai_now().isoformat(),
        "phases": {
            "P0": [spec.table_key for spec in table_specs if spec.phase == "P0"],
            "P1": [spec.table_key for spec in table_specs if spec.phase == "P1"],
        },
        "tables": [
            {
                "table_key": spec.table_key,
                "display_name": spec.display_name,
                "domain": spec.domain,
                "purpose": spec.purpose,
                "expected_rows": spec.expected_rows,
                "plugin_key": spec.plugin_key,
                "phase": spec.phase,
                "default_view_name": spec.default_view_name,
                "field_count": spec.field_count,
                "fields": [
                    {
                        "field_name": field.name,
                        "field_type": field.source_type,
                        "description": field.description,
                        "default": field.default,
                        "required": field.required,
                        "options": list(field.options),
                    }
                    for field in spec.field_specs
                ],
            }
            for spec in table_specs
        ],
    }


def build_view_contract(table_specs: list[TableSpec]) -> dict[str, Any]:
    table_index = {spec.table_key: spec for spec in table_specs}
    return {
        "task_id": TASK_ID,
        "base_name": TARGET_BASE_NAME,
        "created_at": shanghai_now().isoformat(),
        "views": [
            {
                "table_key": view.table_key,
                "table_display_name": table_index[view.table_key].display_name,
                "view_name": view.view_name,
                "view_type": view.view_type,
                "filter_text": view.filter_text,
                "group_sort_text": view.group_sort_text,
                "service_scene": view.service_scene,
            }
            for spec in table_specs
            for view in spec.extra_views
        ],
    }


def _feishu_primary_field_payload(field: FieldSpec) -> dict[str, Any]:
    payload = {"field_name": field.name, "type": field.api_type}
    if field.api_type == 2:
        payload["property"] = {"formatter": number_formatter(field.name)}
    elif field.api_type == 3 or field.api_type == 4:
        payload["property"] = {
            "options": [
                {"name": option, "color": option_color(option, index)}
                for index, option in enumerate(field.options)
            ]
        }
    elif field.api_type == 5:
        payload["property"] = {
            "date_formatter": date_formatter(field.name),
            "auto_fill": auto_fill_date(field.name),
        }
    elif field.api_type in {7, 15, 1}:
        pass
    else:  # pragma: no cover - exhaustive guard
        raise ValueError(f"Unsupported field type id: {field.api_type}")
    return payload


def _normalize_table_name(raw: str | None) -> str:
    return str(raw or "").strip()


def _normalize_view_name(raw: str | None) -> str:
    return str(raw or "").strip()


def _field_meta(field: dict[str, Any]) -> dict[str, Any]:
    return {
        "field_id": str(field.get("field_id") or "").strip(),
        "field_name": str(field.get("field_name") or field.get("name") or "").strip(),
        "type": int(field.get("type") or 0),
        "property": field.get("property"),
    }


def _options_from_meta(meta: dict[str, Any]) -> list[str]:
    property_value = meta.get("property") or {}
    options = property_value.get("options") or []
    return [str(item.get("name") or "").strip() for item in options if str(item.get("name") or "").strip()]


def _get_tables_by_name(api: FeishuBitableAPI, app_token: str) -> dict[str, dict[str, Any]]:
    tables = call_feishu(api, f"/open-apis/bitable/v1/apps/{app_token}/tables?page_size=500").get("data", {}).get("items", [])
    named: dict[str, dict[str, Any]] = {}
    for table in tables:
        name = _normalize_table_name(table.get("name") or table.get("table_name"))
        if name:
            named[name] = table
    return named


def _list_fields(api: FeishuBitableAPI, app_token: str, table_id: str) -> list[dict[str, Any]]:
    response = call_feishu(api, f"/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/fields?page_size=500")
    return response.get("data", {}).get("items", [])


def _list_views(api: FeishuBitableAPI, app_token: str, table_id: str) -> list[dict[str, Any]]:
    response = call_feishu(api, f"/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/views?page_size=500")
    return response.get("data", {}).get("items", [])


def _delete_field(api: FeishuBitableAPI, app_token: str, table_id: str, field_id: str) -> None:
    call_feishu(
        api,
        f"/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/fields/{field_id}",
        method="DELETE",
    )


def _create_table(api: FeishuBitableAPI, app_token: str, spec: TableSpec) -> dict[str, Any]:
    fields = [_feishu_primary_field_payload(spec.field_specs[0])] if spec.field_specs else []
    response = call_feishu(
        api,
        f"/open-apis/bitable/v1/apps/{app_token}/tables",
        method="POST",
        payload={
            "table": {
                "name": spec.table_key,
                "default_view_name": spec.default_view_name,
                "fields": fields,
            }
        },
    )
    return response.get("data", {}).get("table", {}) or response.get("data", {}) or {}


def _rename_base(api: FeishuBitableAPI, app_token: str, name: str) -> dict[str, Any]:
    response = call_feishu(
        api,
        f"/open-apis/bitable/v1/apps/{app_token}",
        method="PUT",
        payload={"name": name},
    )
    return response.get("data", {}).get("app", {}) or response.get("data", {}) or {}


def _rename_table(api: FeishuBitableAPI, app_token: str, table_id: str, name: str) -> dict[str, Any]:
    response = call_feishu(
        api,
        f"/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}",
        method="PATCH",
        payload={"name": name},
    )
    return response.get("data", {}) or {}


def _create_field(api: FeishuBitableAPI, app_token: str, table_id: str, payload: dict[str, Any]) -> dict[str, Any]:
    response = call_feishu(
        api,
        f"/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/fields",
        method="POST",
        payload=payload,
    )
    return response.get("data", {}).get("field", {}) or response.get("data", {}) or {}


def _update_field(api: FeishuBitableAPI, app_token: str, table_id: str, field_id: str, payload: dict[str, Any]) -> dict[str, Any]:
    response = call_feishu(
        api,
        f"/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/fields/{field_id}",
        method="PUT",
        payload=payload,
    )
    return response.get("data", {}).get("field", {}) or response.get("data", {}) or {}


def _create_view(api: FeishuBitableAPI, app_token: str, table_id: str, view_name: str, view_type: str) -> dict[str, Any]:
    response = call_feishu(
        api,
        f"/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/views",
        method="POST",
        payload={"view_name": view_name, "view_type": view_type},
    )
    return response.get("data", {}).get("view", {}) or response.get("data", {}) or {}


def _update_view(api: FeishuBitableAPI, app_token: str, table_id: str, view_id: str, payload: dict[str, Any]) -> dict[str, Any]:
    response = call_feishu(
        api,
        f"/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/views/{view_id}",
        method="PATCH",
        payload=payload,
    )
    return response.get("data", {}).get("view", {}) or response.get("data", {}) or {}


def _get_view(api: FeishuBitableAPI, app_token: str, table_id: str, view_id: str) -> dict[str, Any]:
    response = call_feishu(api, f"/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/views/{view_id}")
    return response.get("data", {}).get("view", {}) or response.get("data", {}) or {}


def _field_value_for_filter(field: FieldSpec, raw_value: str) -> str:
    kind = field_type_kind(field.source_type)
    if kind == "checkbox":
        value = raw_value.strip().lower() in {"1", "true", "yes", "y", "是"}
        return json.dumps([value], ensure_ascii=False)
    if kind == "number":
        try:
            if "." in raw_value:
                return json.dumps([float(raw_value)], ensure_ascii=False)
            return json.dumps([int(raw_value)], ensure_ascii=False)
        except ValueError:
            return json.dumps([raw_value], ensure_ascii=False)
    return json.dumps([raw_value], ensure_ascii=False)


def parse_view_filter_specs(filter_text: str, field_specs: dict[str, FieldSpec]) -> list[dict[str, Any]]:
    text = filter_text.strip()
    if not text or text == "无":
        return []
    chunks = [chunk.strip() for chunk in re.split(r"[;；]\s*", text) if chunk.strip()]
    if not chunks:
        chunks = [text]
    conditions: list[dict[str, Any]] = []
    for chunk in chunks:
        if "=" in chunk:
            field_name, raw_value = chunk.split("=", 1)
            operator = "is"
        elif "!=" in chunk:
            field_name, raw_value = chunk.split("!=", 1)
            operator = "isNot"
        else:
            continue
        field_name = field_name.strip()
        raw_value = raw_value.strip()
        resolved_field_name = field_name
        field = field_specs.get(field_name)
        if field is None:
            alias = VIEW_FILTER_FIELD_ALIASES.get(field_name)
            if alias:
                resolved_field_name = alias
                field = field_specs.get(alias)
        if field is None:
            raise ValueError(f"View filter references unknown field: {field_name}")
        conditions.append(
            {
                "field_name": resolved_field_name,
                "field_type": field.api_type,
                "field_id": None,
                "operator": operator,
                "value": _field_value_for_filter(field, raw_value),
            }
        )
    return conditions


def _build_filter_property(
    conditions: list[dict[str, Any]],
    field_id_by_name: dict[str, str],
) -> dict[str, Any]:
    built_conditions: list[dict[str, Any]] = []
    for condition in conditions:
        field_name = condition["field_name"]
        field_id = field_id_by_name.get(field_name)
        if not field_id:
            raise ValueError(f"Filter field missing in live base: {field_name}")
        built_conditions.append(
            {
                "field_id": field_id,
                "operator": condition["operator"],
                "value": condition["value"],
            }
        )
    return {
        "conditions": built_conditions,
        "conjunction": "and",
    }


def ensure_table(
    api: FeishuBitableAPI,
    app_token: str,
    spec: TableSpec,
    *,
    apply: bool,
    table_lookup: dict[str, dict[str, Any]],
) -> tuple[TableLiveState, dict[str, Any]]:
    current = table_lookup.get(spec.table_key)
    legacy = table_lookup.get("数据表") if spec.table_key == "T01_runtime_control" else None
    created = False
    renamed = False

    if current is None and legacy is not None:
        current = legacy
        if apply:
            _rename_table(api, app_token, str(current.get("table_id") or ""), spec.table_key)
            renamed = True
        current = {**current, "name": spec.table_key}
        table_lookup[spec.table_key] = current
    elif current is None:
        if apply:
            current = _create_table(api, app_token, spec)
            created = True
            table_lookup[spec.table_key] = current
        else:
            current = {"table_id": "", "name": spec.table_key}
    else:
        if _normalize_table_name(current.get("name") or current.get("table_name")) != spec.table_key and apply:
            _rename_table(api, app_token, str(current.get("table_id") or ""), spec.table_key)
            renamed = True
        current = {**current, "name": spec.table_key}

    table_id = str(current.get("table_id") or "").strip()
    if not table_id:
        table_id = ""

    field_meta_by_name: dict[str, dict[str, Any]] = {}
    field_ids_by_name: dict[str, str] = {}
    if table_id:
        for item in _list_fields(api, app_token, table_id):
            meta = _field_meta(item)
            if meta["field_name"]:
                field_meta_by_name[meta["field_name"]] = meta
                field_ids_by_name[meta["field_name"]] = meta["field_id"]

    return (
        TableLiveState(
            table_key=spec.table_key,
            table_id=table_id,
            live_name=spec.table_key,
            field_ids_by_name=field_ids_by_name,
            field_meta_by_name=field_meta_by_name,
            views_by_name={},
            created=created,
            renamed=renamed,
        ),
        current,
    )


def ensure_primary_field(
    api: FeishuBitableAPI,
    app_token: str,
    table: TableLiveState,
    spec: TableSpec,
    *,
    apply: bool,
) -> None:
    if not table.table_id:
        return
    primary = spec.field_specs[0]
    current_fields = table.field_meta_by_name
    if primary.name not in current_fields and current_fields:
        default_meta = next(iter(current_fields.values()))
        if default_meta["field_name"] != primary.name and apply:
            _update_field(
                api,
                app_token,
                table.table_id,
                default_meta["field_id"],
                {"field_name": primary.name, "type": default_meta["type"]},
            )
            table.created_fields.append(primary.name)
            table.field_ids_by_name = {
                **{primary.name: default_meta["field_id"]},
                **{name: field_id for name, field_id in table.field_ids_by_name.items() if name != default_meta["field_name"]},
            }
            refreshed = {}
            refreshed_ids = {}
            for item in _list_fields(api, app_token, table.table_id):
                meta = _field_meta(item)
                if meta["field_name"]:
                    refreshed[meta["field_name"]] = meta
                    refreshed_ids[meta["field_name"]] = meta["field_id"]
            table.field_meta_by_name = refreshed
            table.field_ids_by_name = refreshed_ids


def ensure_field(
    api: FeishuBitableAPI,
    app_token: str,
    table: TableLiveState,
    field_spec: FieldSpec,
    *,
    apply: bool,
) -> dict[str, Any]:
    current = table.field_meta_by_name.get(field_spec.name)
    desired_payload = _feishu_primary_field_payload(field_spec)
    if current:
        current_options = _options_from_meta(current)
        desired_options = list(field_spec.options)
        desired_type = field_spec.api_type
        needs_update = current["type"] != desired_type
        if desired_type in {3, 4} and current_options != desired_options:
            needs_update = True
        if desired_type == 2 and current.get("property", {}).get("formatter") != desired_payload.get("property", {}).get("formatter"):
            needs_update = True
        if desired_type == 5:
            current_property = current.get("property") or {}
            desired_property = desired_payload.get("property") or {}
            if current_property.get("date_formatter") != desired_property.get("date_formatter") or current_property.get("auto_fill") != desired_property.get("auto_fill"):
                needs_update = True
        if needs_update and apply:
            updated = _update_field(api, app_token, table.table_id, current["field_id"], desired_payload)
            table.updated_fields.append(field_spec.name)
            refreshed = _field_meta(updated) if updated else None
            if refreshed and refreshed["field_name"]:
                table.field_meta_by_name[refreshed["field_name"]] = refreshed
                table.field_ids_by_name[refreshed["field_name"]] = refreshed["field_id"]
            else:
                table.field_meta_by_name = {
                    _field_meta(item)["field_name"]: _field_meta(item)
                    for item in _list_fields(api, app_token, table.table_id)
                }
                table.field_ids_by_name = {
                    name: meta["field_id"]
                    for name, meta in table.field_meta_by_name.items()
                }
            return table.field_meta_by_name[field_spec.name]
        return current

    if not apply:
        return {
            "field_id": "",
            "field_name": field_spec.name,
            "type": field_spec.api_type,
            "property": desired_payload.get("property"),
        }

    created = _create_field(api, app_token, table.table_id, desired_payload)
    table.created_fields.append(field_spec.name)
    refreshed = _field_meta(created) if created else None
    if refreshed and refreshed["field_name"]:
        table.field_meta_by_name[refreshed["field_name"]] = refreshed
        table.field_ids_by_name[refreshed["field_name"]] = refreshed["field_id"]
        return refreshed

    table.field_meta_by_name = {
        _field_meta(item)["field_name"]: _field_meta(item)
        for item in _list_fields(api, app_token, table.table_id)
    }
    table.field_ids_by_name = {name: meta["field_id"] for name, meta in table.field_meta_by_name.items()}
    return table.field_meta_by_name[field_spec.name]


def ensure_default_view(
    api: FeishuBitableAPI,
    app_token: str,
    table: TableLiveState,
    spec: TableSpec,
    *,
    apply: bool,
) -> dict[str, Any]:
    if not table.table_id:
        return {"view_id": "", "view_name": spec.default_view_name, "view_type": "grid"}
    views = _list_views(api, app_token, table.table_id)
    by_name = {_normalize_view_name(view.get("view_name") or view.get("name")): view for view in views}
    table.views_by_name = by_name
    if spec.default_view_name in by_name:
        return by_name[spec.default_view_name]
    if not views:
        if apply:
            created = _create_view(api, app_token, table.table_id, spec.default_view_name, "grid")
            table.created_views.append(spec.default_view_name)
            table.views_by_name[spec.default_view_name] = created
            return created
        return {"view_id": "", "view_name": spec.default_view_name, "view_type": "grid"}
    default_view = views[0]
    default_name = _normalize_view_name(default_view.get("view_name") or default_view.get("name"))
    if default_name != spec.default_view_name and apply:
        updated = _update_view(
            api,
            app_token,
            table.table_id,
            str(default_view.get("view_id") or "").strip(),
            {"view_name": spec.default_view_name},
        )
        table.default_view_renamed = True
        refreshed = _get_view(api, app_token, table.table_id, str(default_view.get("view_id") or "").strip())
        if refreshed:
            table.views_by_name[spec.default_view_name] = refreshed
            return refreshed
        if updated:
            table.views_by_name[spec.default_view_name] = updated
            return updated
    return default_view


def ensure_additional_views(
    api: FeishuBitableAPI,
    app_token: str,
    table: TableLiveState,
    spec: TableSpec,
    *,
    apply: bool,
) -> list[dict[str, Any]]:
    if not table.table_id:
        return []
    view_specs = {field.name: field for field in spec.field_specs}
    current_views = _list_views(api, app_token, table.table_id)
    by_name = {_normalize_view_name(view.get("view_name") or view.get("name")): view for view in current_views}
    table.views_by_name = by_name
    created_or_patched: list[dict[str, Any]] = []
    for view_spec in spec.extra_views:
        current = by_name.get(view_spec.view_name)
        if current is None:
            if not apply:
                created_or_patched.append(
                    {
                        "view_name": view_spec.view_name,
                        "view_type": view_spec.view_type,
                        "filter_text": view_spec.filter_text,
                        "group_sort_text": view_spec.group_sort_text,
                        "status": "planned_create",
                    }
                )
                continue
            created = _create_view(api, app_token, table.table_id, view_spec.view_name, view_spec.view_type)
            table.created_views.append(view_spec.view_name)
            by_name[view_spec.view_name] = created
            current = created
        if current.get("view_type") and current.get("view_type") != view_spec.view_type:
            raise RuntimeError(
                f"View type mismatch for {spec.table_key}.{view_spec.view_name}: "
                f"expected {view_spec.view_type}, got {current.get('view_type')}"
            )

        filter_conditions = parse_view_filter_specs(view_spec.filter_text, view_specs)
        if filter_conditions and apply:
            property_payload = _build_filter_property(filter_conditions, table.field_ids_by_name)
            patched = _update_view(
                api,
                app_token,
                table.table_id,
                str(current.get("view_id") or "").strip(),
                {"view_name": view_spec.view_name, "property": property_payload},
            )
            table.patched_views.append(view_spec.view_name)
            if patched:
                by_name[view_spec.view_name] = patched
                current = patched
            _get_view(api, app_token, table.table_id, str(current.get("view_id") or "").strip())
        created_or_patched.append(
            {
                "view_name": view_spec.view_name,
                "view_type": view_spec.view_type,
                "filter_text": view_spec.filter_text,
                "group_sort_text": view_spec.group_sort_text,
                "status": "applied" if apply else "planned",
            }
        )
    table.views_by_name = by_name
    return created_or_patched


def ensure_table_package(
    api: FeishuBitableAPI,
    app_token: str,
    spec: TableSpec,
    *,
    apply: bool,
    table_lookup: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    state, live_table = ensure_table(api, app_token, spec, apply=apply, table_lookup=table_lookup)
    if live_table.get("table_id"):
        state.table_id = str(live_table.get("table_id") or "").strip()
    if state.table_id and apply:
        actual = _normalize_table_name(live_table.get("name") or live_table.get("table_name") or spec.table_key)
        if actual != spec.table_key:
            state.renamed = True
    if state.table_id:
        if not state.field_meta_by_name and apply:
            state.field_meta_by_name = {
                _field_meta(item)["field_name"]: _field_meta(item) for item in _list_fields(api, app_token, state.table_id)
            }
            state.field_ids_by_name = {name: meta["field_id"] for name, meta in state.field_meta_by_name.items()}
        ensure_primary_field(api, app_token, state, spec, apply=apply)
        for field_spec in spec.field_specs:
            ensure_field(api, app_token, state, field_spec, apply=apply)
        if apply:
            expected_names = {field.name for field in spec.field_specs}
            current_fields = _list_fields(api, app_token, state.table_id)
            extra_fields = [
                item
                for item in current_fields
                if str(item.get("field_name") or item.get("name") or "").strip() not in expected_names
                and not item.get("is_primary")
            ]
            for item in extra_fields:
                field_id = str(item.get("field_id") or "").strip()
                if field_id:
                    _delete_field(api, app_token, state.table_id, field_id)
            if extra_fields:
                state.field_meta_by_name = {
                    _field_meta(item)["field_name"]: _field_meta(item) for item in _list_fields(api, app_token, state.table_id)
                }
                state.field_ids_by_name = {name: meta["field_id"] for name, meta in state.field_meta_by_name.items()}
        ensure_default_view(api, app_token, state, spec, apply=apply)
        ensure_additional_views(api, app_token, state, spec, apply=apply)

    if state.table_id:
        final_fields = _list_fields(api, app_token, state.table_id)
        final_views = _list_views(api, app_token, state.table_id)
    else:
        final_fields = []
        final_views = []

    final_field_names = [str(item.get("field_name") or item.get("name") or "").strip() for item in final_fields if str(item.get("field_name") or item.get("name") or "").strip()]
    expected_field_names = [field.name for field in spec.field_specs]
    if apply and final_field_names != expected_field_names:
        raise RuntimeError(
            f"Field mismatch for {spec.table_key}: expected {expected_field_names}, got {final_field_names}"
        )

    final_view_names = [str(item.get("view_name") or item.get("name") or "").strip() for item in final_views if str(item.get("view_name") or item.get("name") or "").strip()]
    expected_view_names = [spec.default_view_name, *[view.view_name for view in spec.extra_views]]
    for view_name in expected_view_names:
        if view_name not in final_view_names:
            if apply:
                raise RuntimeError(f"View mismatch for {spec.table_key}: missing {view_name}")

    view_names = [spec.default_view_name, *[view.view_name for view in spec.extra_views]]
    registry_entry = {
        "table_id": state.table_id,
        "field_count": len(final_fields) if final_fields else spec.field_count,
        "views": [name for name in view_names if name],
    }
    return {
        "table_key": spec.table_key,
        "display_name": spec.display_name,
        "domain": spec.domain,
        "purpose": spec.purpose,
        "phase": spec.phase,
        "table_id": state.table_id,
        "created": state.created,
        "renamed": state.renamed,
        "default_view_renamed": state.default_view_renamed,
        "created_fields": state.created_fields,
        "updated_fields": state.updated_fields,
        "created_views": state.created_views,
        "patched_views": state.patched_views,
        "field_count": registry_entry["field_count"],
        "views": registry_entry["views"],
        "field_names": final_field_names or expected_field_names,
        "view_names": final_view_names or view_names,
    }


def build_target(api: FeishuBitableAPI, app_token: str, table_specs: list[TableSpec], *, apply: bool) -> dict[str, Any]:
    table_lookup = _get_tables_by_name(api, app_token)
    base_info: dict[str, Any] = {}
    if apply:
        base_info = _rename_base(api, app_token, TARGET_BASE_NAME)
    else:
        base_info = call_feishu(api, f"/open-apis/bitable/v1/apps/{app_token}").get("data", {}).get("app", {}) or {}

    results: list[dict[str, Any]] = []
    for phase in ("P0", "P1"):
        for spec in [item for item in table_specs if item.phase == phase]:
            results.append(ensure_table_package(api, app_token, spec, apply=apply, table_lookup=table_lookup))

    registry = {
        "base_name": TARGET_BASE_NAME if apply else str(base_info.get("name") or TARGET_BASE_NAME),
        "app_token": app_token,
        "created_at": shanghai_now().isoformat(),
        "tables": {
            item["table_key"]: {
                "table_id": item["table_id"],
                "field_count": item["field_count"],
                "views": item["views"],
            }
            for item in results
        },
    }

    detailed_registry = {
        "task_id": TASK_ID,
        "base_name": registry["base_name"],
        "app_token": app_token,
        "created_at": registry["created_at"],
        "tables": results,
    }

    return {
        "registry": registry,
        "detailed_registry": detailed_registry,
        "base_info": base_info,
        "table_results": results,
    }


def write_artifacts(run_dir: Path, *, payload: dict[str, Any], mode: str) -> None:
    run_dir.mkdir(parents=True, exist_ok=True)
    json_write(run_dir / f"{mode}.json", payload)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Build the PROJ-V5 base from the V2 xlsx spec.")
    parser.add_argument("--link", default=DEFAULT_LINK, help="Feishu wiki/base link for the new base.")
    parser.add_argument("--account-id", default=DEFAULT_ACCOUNT_ID)
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--dry-run", action="store_true")
    mode.add_argument("--apply", action="store_true")
    args = parser.parse_args(argv)

    apply_changes = bool(args.apply)
    run_id = f"adagj-{TASK_ID.lower()}-{now_stamp()}"
    run_dir = ARTIFACT_ROOT / datetime.now().strftime("%Y-%m-%d") / run_id
    run_dir.mkdir(parents=True, exist_ok=True)

    creds = load_feishu_credentials(args.account_id)
    os.environ["FEISHU_APP_ID"] = creds["app_id"]
    os.environ["FEISHU_APP_SECRET"] = creds["app_secret"]
    api = FeishuBitableAPI(creds["app_id"], creds["app_secret"])

    table_specs = parse_field_specs(WORKBOOK_PATH)
    schema_manifest = build_schema_manifest(table_specs)
    view_contract = build_view_contract(table_specs)
    json_write(SCHEMA_MANIFEST_PATH, schema_manifest)
    json_write(VIEW_CONTRACT_PATH, view_contract)

    app_token = api.resolve_app_token(args.link)
    base_before = call_feishu(api, f"/open-apis/bitable/v1/apps/{app_token}").get("data", {}).get("app", {}) or {}

    result_payload: dict[str, Any] = {
        "ok": True,
        "task_id": TASK_ID,
        "mode": "apply" if apply_changes else "dry-run",
        "link": args.link,
        "account_id": args.account_id,
        "app_token": app_token,
        "base_before": base_before,
        "schema_manifest_path": str(SCHEMA_MANIFEST_PATH),
        "view_contract_path": str(VIEW_CONTRACT_PATH),
        "run_dir": str(run_dir),
    }

    try:
        target = build_target(api, app_token, table_specs, apply=apply_changes)
        result_payload.update(
            {
                "status": "schema_applied" if apply_changes else "schema_preview_ready",
                "base_after": target["base_info"],
                "registry_path": str(REGISTRY_PATH),
                "tables": target["table_results"],
                "registry": target["registry"],
            }
        )
        if apply_changes:
            json_write(REGISTRY_PATH, target["registry"])
            json_write(run_dir / "apply.json", result_payload)
        else:
            json_write(run_dir / "dry-run.json", result_payload)
            json_write(PARTIAL_REGISTRY_PATH, target["registry"])
    except Exception as exc:  # pragma: no cover - live API only
        result_payload.update({"ok": False, "status": "failed", "error": str(exc)})
        if apply_changes:
            try:
                json_write(PARTIAL_REGISTRY_PATH, result_payload)
            except Exception:
                pass
            json_write(run_dir / "apply.json", result_payload)
        else:
            json_write(run_dir / "dry-run.json", result_payload)
        print(json.dumps(result_payload, ensure_ascii=False, indent=2))
        return 1

    print(json.dumps(result_payload, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
