#!/usr/bin/env python3
"""Apply TS-YS-STEPA-01 structural data model fixes to the live Feishu base."""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from dashboard.feishu_deploy import FeishuBitableAPI

WORKBOOK_PATH = Path("/Users/liming/Documents/会议纪要/【艾琳项目】/驾驶舱（三期，勿动）.xlsx")
DOCX_PATH = Path("/Users/liming/Downloads/TS-YS-STEPA-01_数据模型结构性补缺 (1).docx")
DEFAULT_APP_TOKEN = "CtltbMK3IaXLX8s0mTVcBw2Anth"
BASE_LINK = f"https://h52xu4gwob.feishu.cn/base/{DEFAULT_APP_TOKEN}"
ARTIFACT_ROOT = REPO_ROOT / "artifacts" / "ai-da-guan-jia" / "runs"
TASK_ID = "TS-YS-STEPA-01"
SHANGHAI_TZ = timezone(timedelta(hours=8))

FIELD_TEXT = 1
FIELD_NUMBER = 2
FIELD_SINGLE_SELECT = 3
FIELD_MULTI_SELECT = 4
FIELD_DATETIME = 5
FIELD_CHECKBOX = 7
FIELD_URL = FIELD_TEXT
FIELD_MULTILINE_TEXT = FIELD_TEXT
FIELD_RELATION = 21


@dataclass(frozen=True)
class FieldSpec:
    name: str
    type: str
    options: tuple[str, ...] = ()
    property: dict[str, Any] = field(default_factory=dict)


@dataclass(frozen=True)
class TableSpec:
    table_name: str
    primary_field: str
    fields: tuple[FieldSpec, ...]
    source_sheet: str | None = None

    def field_map(self) -> dict[str, FieldSpec]:
        return {field.name: field for field in self.fields}

    def create_fields(self) -> list[FieldSpec]:
        return [field for field in self.fields if field.type != "relation"]


@dataclass(frozen=True)
class RelationSpec:
    source_table: str
    source_field: str
    target_table: str
    target_field: str
    multiple: bool = False
    back_field_name: str = ""

    @property
    def link_field_name(self) -> str:
        return f"{self.source_field}_link"


def now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d-%H%M%S")


def shanghai_now() -> datetime:
    return datetime.now(tz=SHANGHAI_TZ)


def as_ms(value: Any) -> int | str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        dt = value if value.tzinfo else value.replace(tzinfo=SHANGHAI_TZ)
        return int(dt.timestamp() * 1000)
    if isinstance(value, date):
        dt = datetime(value.year, value.month, value.day, tzinfo=SHANGHAI_TZ)
        return int(dt.timestamp() * 1000)
    if isinstance(value, (int, float)):
        numeric = int(value)
        return numeric if abs(numeric) >= 10**12 else numeric * 1000
    text = str(value).strip()
    if not text:
        return ""
    if text.isdigit():
        numeric = int(text)
        return numeric if abs(numeric) >= 10**12 else numeric * 1000
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=SHANGHAI_TZ)
        return int(parsed.timestamp() * 1000)
    except ValueError:
        return text


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def write_json(path: Path, payload: Any) -> None:
    ensure_dir(path.parent)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_text(path: Path, text: str) -> None:
    ensure_dir(path.parent)
    path.write_text(text.rstrip() + "\n", encoding="utf-8")


def load_credentials() -> tuple[str, str]:
    app_id = str(os.getenv("FEISHU_APP_ID") or "").strip()
    app_secret = str(os.getenv("FEISHU_APP_SECRET") or "").strip()
    if app_id and app_secret:
        return app_id, app_secret
    try:
        from scripts.create_kangbo_signal_tables import load_feishu_credentials
    except Exception as exc:  # pragma: no cover - fallback path
        raise RuntimeError("Missing FEISHU_APP_ID / FEISHU_APP_SECRET and fallback loader import failed.") from exc
    creds = load_feishu_credentials()
    os.environ["FEISHU_APP_ID"] = creds["app_id"]
    os.environ["FEISHU_APP_SECRET"] = creds["app_secret"]
    return creds["app_id"], creds["app_secret"]


def load_api() -> FeishuBitableAPI:
    app_id, app_secret = load_credentials()
    return FeishuBitableAPI(app_id, app_secret)


def split_tokens(value: Any) -> list[str]:
    if value is None or value == "":
        return []
    if isinstance(value, list):
        tokens: list[str] = []
        for item in value:
            tokens.extend(split_tokens(item))
        return tokens
    text = str(value).strip()
    if not text:
        return []
    if text in {"None", "none", "nan"}:
        return []
    pieces = re.split(r"[,\n，、;/+|]+", text)
    return [piece.strip() for piece in pieces if piece.strip()]


def unique_tokens(*values: Any) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        for token in split_tokens(value):
            if token not in seen:
                seen.add(token)
                ordered.append(token)
    return ordered


def load_sheet_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for row_idx in range(2, sheet.max_row + 1):
        values = [sheet.cell(row_idx, col).value for col in range(1, sheet.max_column + 1)]
        if not any(value not in (None, "") for value in values):
            continue
        row = {str(headers[col]).strip(): values[col] for col in range(len(headers)) if headers[col] not in (None, "")}
        rows.append(row)
    return rows


def load_workbook_rows() -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    rows: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in workbook.sheetnames:
        rows[sheet_name] = load_sheet_rows(WORKBOOK_PATH, sheet_name)
    return rows


def field_payload(field: FieldSpec) -> dict[str, Any]:
    type_map = {
        "text": FIELD_TEXT,
        "number": FIELD_NUMBER,
        "single_select": FIELD_SINGLE_SELECT,
        "multi_select": FIELD_MULTI_SELECT,
        "datetime": FIELD_DATETIME,
        "checkbox": FIELD_CHECKBOX,
        "url": FIELD_URL,
        "multiline_text": FIELD_MULTILINE_TEXT,
        "relation": FIELD_RELATION,
    }
    if field.type not in type_map:
        raise ValueError(f"Unsupported field type: {field.type}")
    payload: dict[str, Any] = {"field_name": field.name, "type": type_map[field.type]}
    if field.type in {"single_select", "multi_select"}:
        payload["property"] = {"options": [{"name": option} for option in field.options]}
    elif field.type == "datetime":
        payload["property"] = {"date_formatter": field.property.get("date_formatter", "yyyy-MM-dd"), "auto_fill": False}
    elif field.type == "relation":
        payload["property"] = dict(field.property)
    elif field.type == "url":
        payload["property"] = {}
    elif field.type == "number" and field.property:
        payload["property"] = dict(field.property)
    return payload


def normalize_record_value(field: FieldSpec, value: Any) -> Any:
    if value is None or value == "":
        if field.type == "multi_select":
            return []
        return ""
    if field.type == "checkbox":
        if isinstance(value, bool):
            return value
        text = str(value).strip()
        return text in {"1", "true", "True", "yes", "y", "是"}
    if field.type == "number":
        if isinstance(value, (int, float)):
            return value
        text = str(value).strip().replace(",", "")
        if not text:
            return ""
        try:
            return int(text)
        except ValueError:
            try:
                return round(float(text), 2)
            except ValueError:
                return text
    if field.type == "datetime":
        return as_ms(value)
    if field.type == "multi_select":
        if isinstance(value, list):
            return [str(item).strip() for item in value if str(item).strip()]
        return split_tokens(value)
    if field.type == "relation":
        if isinstance(value, list):
            return [str(item).strip() for item in value if str(item).strip()]
        tokens = split_tokens(value)
        return tokens if tokens else []
    if field.type in {"text", "multiline_text", "url"}:
        return str(value).strip()
    return value


def normalize_row(row: dict[str, Any], field_map: dict[str, FieldSpec]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for key, value in row.items():
        field = field_map.get(key)
        if not field:
            continue
        normalized_value = normalize_record_value(field, value)
        if normalized_value == "" or normalized_value == [] or normalized_value is None:
            continue
        normalized[key] = normalized_value
    return normalized


def sheet_value(rows: list[dict[str, Any]], key: str) -> list[Any]:
    return [row.get(key) for row in rows]


def collect_options(rows: list[dict[str, Any]], key: str) -> list[str]:
    return unique_tokens(*sheet_value(rows, key))


def shanghai_date(year: int, month: int, day: int, hour: int = 9, minute: int = 0) -> datetime:
    return datetime(year, month, day, hour, minute, tzinfo=SHANGHAI_TZ)


def month_rank(value: Any) -> int:
    text = str(value or "").strip()
    mapping = {"全年": 12, "上半年": 6, "一季度": 3, "二季度": 6, "三季度": 9, "四季度": 12, "1月": 1, "2月": 2, "3月": 3, "4月": 4, "5月": 5, "6月": 6, "7月": 7, "8月": 8, "9月": 9, "10月": 10, "11月": 11, "12月": 12}
    if text in mapping:
        return mapping[text]
    digits = "".join(ch for ch in text if ch.isdigit())
    if digits:
        month = int(digits[:2] if len(digits) > 1 and int(digits[:2]) <= 12 else digits[0])
        return month
    return 0


def project_status_rank(status: str) -> int:
    order = {"在谈": 1, "已签约": 2, "已开工": 3, "建设中": 4, "已投产": 5}
    return order.get(status, 0)


def lead_stage_rank(stage: str) -> int:
    order = {"签约": 1, "落地": 2, "意向": 3, "对接": 4, "储备": 5, "停滞": 6}
    return order.get(stage, 99)


def build_table_specs(ctx: dict[str, Any]) -> tuple[list[TableSpec], list[RelationSpec]]:
    raw = ctx["raw"]
    t1 = raw["T1_企业主数据"]
    t2 = raw["T2_重点项目"]
    t3 = raw["T3_招商跟进"]
    t4 = raw["T4_企业画像"]
    t5 = raw["T5_产业链主数据"]
    t6 = raw["T6_政策主数据"]
    t6a = raw["T6A_政策申报记录"]
    t7 = raw["T7_空间资源"]
    t7a = raw["T7A_空间对接记录"]
    t8 = raw["T8_组织责任人"]

    t1m_fields = (
        FieldSpec("记录ID", "text"),
        FieldSpec("企业", "text"),
        FieldSpec("企业_link", "relation", property={"table_id": "", "multiple": False, "back_field_name": "T1M_企业月度经营_back"}),
        FieldSpec("年份", "number"),
        FieldSpec("月份", "text"),
        FieldSpec("当期产值(万元)", "number"),
        FieldSpec("累计产值(万元)", "number"),
        FieldSpec("去年同期累计(万元)", "number"),
        FieldSpec("目标产值(万元)", "number"),
        FieldSpec("当期税收(万元)", "number"),
        FieldSpec("备注", "multiline_text"),
    )

    t1_fields = (
        FieldSpec("企业ID", "text"),
        FieldSpec("企业名称", "text"),
        FieldSpec("企业简称", "text"),
        FieldSpec("企业类型", "single_select", options=tuple(collect_options(t1, "企业类型"))),
        FieldSpec("所属产业链", "multi_select", options=tuple(unique_tokens(*sheet_value(t1, "所属产业链")))),
        FieldSpec("资质等级", "single_select", options=tuple(collect_options(t1, "资质等级"))),
        FieldSpec("是否规上", "checkbox"),
        FieldSpec("区级考核", "single_select", options=tuple(collect_options(t1, "区级考核"))),
        FieldSpec("入驻载体", "single_select", options=tuple(collect_options(t1, "入驻载体"))),
        FieldSpec("注册地址", "text"),
        FieldSpec("法人代表", "text"),
        FieldSpec("联系人", "text"),
        FieldSpec("联系电话", "text"),
        FieldSpec("企业简介", "multiline_text"),
        FieldSpec("备注", "multiline_text"),
    )

    t2_fields = (
        FieldSpec("项目ID", "text"),
        FieldSpec("项目名称", "text"),
        FieldSpec("产业方向", "single_select", options=tuple(collect_options(t2, "产业方向"))),
        FieldSpec("建设单位", "text"),
        FieldSpec("关联企业", "text"),
        FieldSpec("总投资(万元)", "number"),
        FieldSpec("预期年产值(万元)", "number"),
        FieldSpec("预期年税收(万元)", "number"),
        FieldSpec("项目简介", "multiline_text"),
        FieldSpec("当前状态", "single_select", options=tuple(collect_options(t2, "当前状态"))),
        FieldSpec("建设阶段", "single_select", options=tuple(collect_options(t2, "建设阶段"))),
        FieldSpec("责任人", "text"),
        FieldSpec("所在空间", "text"),
        FieldSpec("计划开工日期", "datetime"),
        FieldSpec("计划投产日期", "datetime"),
        FieldSpec("实际开工日期", "datetime"),
        FieldSpec("实际投产日期", "datetime"),
        FieldSpec("需协调事项", "multiline_text"),
        FieldSpec("最新进展", "multiline_text"),
        FieldSpec("更新日期", "datetime"),
        FieldSpec("备注", "multiline_text"),
        FieldSpec("来源招商ID", "text"),
        FieldSpec("来源招商ID_link", "relation", property={"table_id": "", "multiple": False, "back_field_name": "T2_重点项目_back"}),
        FieldSpec("项目来源", "single_select", options=("招商转化", "存量扩建", "政策引导", "其他")),
        FieldSpec("全周期状态", "single_select", options=("线索→研判", "研判→签约", "签约→建设", "建设→投产", "投产→达产")),
    )

    t2m_fields = (
        FieldSpec("里程碑ID", "text"),
        FieldSpec("项目", "text"),
        FieldSpec("项目_link", "relation", property={"table_id": "", "multiple": False, "back_field_name": "T2M_项目里程碑_back"}),
        FieldSpec("里程碑名称", "text"),
        FieldSpec("计划完成日期", "datetime"),
        FieldSpec("实际完成日期", "datetime"),
        FieldSpec("状态", "single_select", options=tuple(collect_options(raw["T2M_项目里程碑"], "状态"))),
        FieldSpec("卡点说明", "multiline_text"),
        FieldSpec("需协调事项", "multiline_text"),
        FieldSpec("责任人", "text"),
    )

    t3_fields = (
        FieldSpec("招商ID", "text"),
        FieldSpec("企业/项目名称", "text"),
        FieldSpec("阶段", "single_select", options=tuple(collect_options(t3, "阶段"))),
        FieldSpec("产业方向", "single_select", options=tuple(collect_options(t3, "产业方向"))),
        FieldSpec("来源渠道", "single_select", options=tuple(collect_options(t3, "来源渠道"))),
        FieldSpec("活动来源", "single_select", options=tuple(collect_options(t3, "活动来源"))),
        FieldSpec("预期投资(万元)", "number"),
        FieldSpec("预期产值(万元)", "number"),
        FieldSpec("首次对接日期", "datetime"),
        FieldSpec("最新进展日期", "datetime"),
        FieldSpec("跟进人", "text"),
        FieldSpec("进展记录", "multiline_text"),
        FieldSpec("是否云产业园", "checkbox"),
        FieldSpec("是否两区项目", "checkbox"),
        FieldSpec("签约日期", "datetime"),
        FieldSpec("落地日期", "datetime"),
        FieldSpec("备注", "multiline_text"),
        FieldSpec("转化项目ID", "text"),
        FieldSpec("转化项目ID_link", "relation", property={"table_id": "", "multiple": False, "back_field_name": "T3_招商跟进_back"}),
        FieldSpec("转化日期", "datetime"),
        FieldSpec("转化状态", "single_select", options=("未转化", "已转化", "放弃")),
        FieldSpec("流转阶段", "single_select", options=("线索", "研判", "座谈", "考察", "签约", "建设", "投产")),
        FieldSpec("当前卡点", "multiline_text"),
    )

    t4_fields = (
        FieldSpec("企业", "text"),
        FieldSpec("最新年度产值(万元)", "number"),
        FieldSpec("最新年度税收(万元)", "number"),
        FieldSpec("研发投入(万元)", "number"),
        FieldSpec("员工人数", "number"),
        FieldSpec("成长状态", "single_select", options=tuple(collect_options(t4, "成长状态"))),
        FieldSpec("核心产品", "multiline_text"),
        FieldSpec("技术优势", "multiline_text"),
        FieldSpec("政策需求", "multiline_text"),
        FieldSpec("适配政策", "multi_select", options=tuple(unique_tokens(*[row.get("政策名称") for row in t6]))),
        FieldSpec("入驻空间", "single_select", options=tuple(collect_options(t4, "入驻空间"))),
        FieldSpec("成长预警", "single_select", options=("高速增长", "稳定增长", "持平", "下滑", "停产", "预警")),
        FieldSpec("关联项目", "text"),
        FieldSpec("关联项目_link", "relation", property={"table_id": "", "multiple": True, "back_field_name": "T4_企业画像_项目_back"}),
        FieldSpec("关联招商", "text"),
        FieldSpec("关联招商_link", "relation", property={"table_id": "", "multiple": True, "back_field_name": "T4_企业画像_招商_back"}),
        FieldSpec("适配政策_link", "relation", property={"table_id": "", "multiple": True, "back_field_name": "T4_企业画像_政策_back"}),
    )

    t5_fields = (
        FieldSpec("链ID", "text"),
        FieldSpec("链名称", "text"),
        FieldSpec("链主企业", "text"),
        FieldSpec("链长", "text"),
        FieldSpec("工作专班", "text"),
        FieldSpec("目标产值(亿元)", "number"),
        FieldSpec("当前产值(亿元)", "number"),
        FieldSpec("上游已有环节", "multiline_text"),
        FieldSpec("上游缺失环节", "multiline_text"),
        FieldSpec("中游已有环节", "multiline_text"),
        FieldSpec("中游缺失环节", "multiline_text"),
        FieldSpec("下游已有环节", "multiline_text"),
        FieldSpec("下游缺失环节", "multiline_text"),
        FieldSpec("重点招商方向", "multiline_text"),
        FieldSpec("在建补链项目", "multiline_text"),
        FieldSpec("链上重点企业", "multiline_text"),
        FieldSpec("主要风险", "multiline_text"),
        FieldSpec("链上企业数", "number"),
        FieldSpec("链上规上企业数", "number"),
        FieldSpec("招商缺口分析", "multiline_text"),
    )

    t6_fields = (
        FieldSpec("政策ID", "text"),
        FieldSpec("政策名称", "text"),
        FieldSpec("级别", "single_select", options=tuple(collect_options(t6, "级别"))),
        FieldSpec("发布部门", "text"),
        FieldSpec("适用领域", "multi_select", options=tuple(unique_tokens(*sheet_value(t6, "适用领域")))),
        FieldSpec("关键词标签", "multi_select", options=tuple(unique_tokens(*sheet_value(t6, "关键词标签")))),
        FieldSpec("最高额度(万元)", "number"),
        FieldSpec("申报截止日", "datetime"),
        FieldSpec("原文链接", "url"),
        FieldSpec("状态", "single_select", options=("有效", "即将到期")),
        FieldSpec("政策摘要", "multiline_text"),
        FieldSpec("申报条件", "multiline_text"),
        FieldSpec("备注", "multiline_text"),
    )

    t6a_fields = (
        FieldSpec("申报ID", "text"),
        FieldSpec("政策", "text"),
        FieldSpec("政策_link", "relation", property={"table_id": "", "multiple": False, "back_field_name": "T6A_政策申报记录_政策_back"}),
        FieldSpec("申报企业", "text"),
        FieldSpec("关联企业ID", "text"),
        FieldSpec("关联企业ID_link", "relation", property={"table_id": "", "multiple": False, "back_field_name": "T6A_政策申报记录_企业_back"}),
        FieldSpec("申报日期", "datetime"),
        FieldSpec("申报金额(万元)", "number"),
        FieldSpec("审批状态", "single_select", options=tuple(collect_options(t6a, "审批状态"))),
        FieldSpec("兑现金额(万元)", "number"),
        FieldSpec("兑现日期", "datetime"),
        FieldSpec("备注", "multiline_text"),
        FieldSpec("企业申报前年产值(万元)", "number"),
        FieldSpec("企业最新年产值(万元)", "number"),
        FieldSpec("产值增量(万元)", "number"),
        FieldSpec("企业申报前年税收(万元)", "number"),
        FieldSpec("企业最新年税收(万元)", "number"),
        FieldSpec("税收增量(万元)", "number"),
        FieldSpec("投产比", "number"),
        FieldSpec("政策效果评级", "single_select", options=("高效", "有效", "低效", "待评估")),
        FieldSpec("效果说明", "multiline_text"),
    )

    t6b_fields = (
        FieldSpec("模拟ID", "text"),
        FieldSpec("模拟场景名称", "text"),
        FieldSpec("目标产业", "single_select", options=("电子信息", "石墨烯", "氢能", "高性能膜", "合成材料", "智能制造", "安全应急")),
        FieldSpec("资金池规模(万元)", "number"),
        FieldSpec("扶持企业数", "number"),
        FieldSpec("预期新增产值(万元)", "number"),
        FieldSpec("预期新增税收(万元)", "number"),
        FieldSpec("预期新增就业(人)", "number"),
        FieldSpec("预期投产比", "number"),
        FieldSpec("模拟假设", "multiline_text"),
        FieldSpec("创建人", "text"),
        FieldSpec("创建日期", "datetime"),
    )

    t7_fields = (
        FieldSpec("空间ID", "text"),
        FieldSpec("名称/位置", "text"),
        FieldSpec("类型", "single_select", options=tuple(collect_options(t7, "类型"))),
        FieldSpec("权属", "single_select", options=tuple(collect_options(t7, "权属"))),
        FieldSpec("面积(m²)", "number"),
        FieldSpec("配套设施", "multiline_text"),
        FieldSpec("适配产业", "multi_select", options=tuple(unique_tokens(*sheet_value(t7, "适配产业")))),
        FieldSpec("状态", "single_select", options=tuple(collect_options(t7, "状态"))),
        FieldSpec("意向/签约企业", "text"),
        FieldSpec("地图平台链接", "url"),
        FieldSpec("备注", "multiline_text"),
    )

    t7a_fields = (
        FieldSpec("对接ID", "text"),
        FieldSpec("空间", "text"),
        FieldSpec("空间_link", "relation", property={"table_id": "", "multiple": False, "back_field_name": "T7A_空间对接记录_back"}),
        FieldSpec("企业名称", "text"),
        FieldSpec("对接日期", "datetime"),
        FieldSpec("进展", "multiline_text"),
        FieldSpec("签约状态", "single_select", options=tuple(collect_options(t7a, "签约状态"))),
    )

    t8_fields = (
        FieldSpec("人员ID", "text"),
        FieldSpec("姓名", "text"),
        FieldSpec("部门", "text"),
        FieldSpec("职务", "text"),
        FieldSpec("角色", "multi_select", options=tuple(unique_tokens(*sheet_value(t8, "角色")))),
        FieldSpec("负责模块", "multi_select", options=tuple(unique_tokens(*sheet_value(t8, "负责模块")))),
        FieldSpec("联系方式", "text"),
    )

    t9_fields = (
        FieldSpec("预警ID", "text"),
        FieldSpec("预警类型", "single_select", options=("项目逾期", "招商停滞", "政策待办", "空间闲置", "协同阻塞", "其他")),
        FieldSpec("来源表", "single_select", options=("T2_重点项目", "T3_招商跟进", "T6A_政策申报记录", "T7_空间资源", "T10A_项目协同记录")),
        FieldSpec("触发指标", "text"),
        FieldSpec("触发值", "text"),
        FieldSpec("灯色", "single_select", options=("红灯", "黄灯")),
        FieldSpec("关联记录", "text"),
        FieldSpec("责任人", "text"),
        FieldSpec("处置状态", "single_select", options=("待处理", "处理中", "已关闭")),
        FieldSpec("督办时限", "datetime"),
        FieldSpec("处置记录", "multiline_text"),
        FieldSpec("创建时间", "datetime"),
        FieldSpec("关闭时间", "datetime"),
    )

    t10_fields = (
        FieldSpec("资源ID", "text"),
        FieldSpec("资源名称", "text"),
        FieldSpec("资源圈层", "single_select", options=("内部部门", "委办局", "生态伙伴")),
        FieldSpec("资源类型", "single_select", options=("审批", "评估", "财税", "法律", "科技", "政策", "招商", "其他")),
        FieldSpec("联系人", "text"),
        FieldSpec("联系方式", "text"),
        FieldSpec("擅长服务", "multiline_text"),
        FieldSpec("关联产业", "multi_select", options=("电子信息", "石墨烯", "氢能", "高性能膜", "合成材料", "智能制造", "安全应急", "绿色油品", "其他")),
        FieldSpec("备注", "multiline_text"),
    )

    t10a_fields = (
        FieldSpec("记录ID", "text"),
        FieldSpec("项目", "text"),
        FieldSpec("项目_link", "relation", property={"table_id": "", "multiple": False, "back_field_name": "T10A_项目协同记录_项目_back"}),
        FieldSpec("协同资源", "text"),
        FieldSpec("协同资源_link", "relation", property={"table_id": "", "multiple": False, "back_field_name": "T10A_项目协同记录_资源_back"}),
        FieldSpec("协同环节", "single_select", options=("立项", "环评", "安评", "建设许可", "设备采购", "人才引进", "资金对接", "验收", "其他")),
        FieldSpec("协同状态", "single_select", options=("待启动", "进行中", "已完成", "阻塞")),
        FieldSpec("开始日期", "datetime"),
        FieldSpec("完成日期", "datetime"),
        FieldSpec("卡点说明", "multiline_text"),
        FieldSpec("责任人", "text"),
    )

    t15_fields = (
        FieldSpec("KPI名称", "text"),
        FieldSpec("计算方式", "multiline_text"),
        FieldSpec("数据源", "text"),
        FieldSpec("当前值", "number"),
        FieldSpec("备注", "multiline_text"),
    )

    tables = [
        TableSpec("T1M_企业月度经营", "记录ID", t1m_fields, "T1M_企业月度经营"),
        TableSpec("T1_企业主数据", "企业ID", t1_fields, "T1_企业主数据"),
        TableSpec("T2_重点项目", "项目ID", t2_fields, "T2_重点项目"),
        TableSpec("T2M_项目里程碑", "里程碑ID", t2m_fields, "T2M_项目里程碑"),
        TableSpec("T3_招商跟进", "招商ID", t3_fields, "T3_招商跟进"),
        TableSpec("T4_企业画像", "企业", t4_fields, "T4_企业画像"),
        TableSpec("T5_产业链主数据", "链ID", t5_fields, "T5_产业链主数据"),
        TableSpec("T6_政策主数据", "政策ID", t6_fields, "T6_政策主数据"),
        TableSpec("T6A_政策申报记录", "申报ID", t6a_fields, "T6A_政策申报记录"),
        TableSpec("T6B_政策模拟", "模拟ID", t6b_fields, None),
        TableSpec("T7_空间资源", "空间ID", t7_fields, "T7_空间资源"),
        TableSpec("T7A_空间对接记录", "对接ID", t7a_fields, "T7A_空间对接记录"),
        TableSpec("T8_组织责任人", "人员ID", t8_fields, "T8_组织责任人"),
        TableSpec("T9_预警事件", "预警ID", t9_fields, None),
        TableSpec("T10_协同资源", "资源ID", t10_fields, None),
        TableSpec("T10A_项目协同记录", "记录ID", t10a_fields, None),
        TableSpec("T15_KPI汇总", "KPI名称", t15_fields, "T15_KPI汇总"),
    ]

    relations = [
        RelationSpec("T1M_企业月度经营", "企业", "T1_企业主数据", "企业名称", False, "T1M_企业月度经营_back"),
        RelationSpec("T2_重点项目", "来源招商ID", "T3_招商跟进", "招商ID", False, "T2_重点项目_back"),
        RelationSpec("T2M_项目里程碑", "项目", "T2_重点项目", "项目名称", False, "T2M_项目里程碑_back"),
        RelationSpec("T3_招商跟进", "转化项目ID", "T2_重点项目", "项目ID", False, "T3_招商跟进_back"),
        RelationSpec("T4_企业画像", "关联项目", "T2_重点项目", "项目ID", True, "T4_企业画像_项目_back"),
        RelationSpec("T4_企业画像", "关联招商", "T3_招商跟进", "招商ID", True, "T4_企业画像_招商_back"),
        RelationSpec("T4_企业画像", "适配政策", "T6_政策主数据", "政策名称", True, "T4_企业画像_政策_back"),
        RelationSpec("T6A_政策申报记录", "政策", "T6_政策主数据", "政策名称", False, "T6A_政策申报记录_政策_back"),
        RelationSpec("T6A_政策申报记录", "关联企业ID", "T1_企业主数据", "企业ID", False, "T6A_政策申报记录_企业_back"),
        RelationSpec("T7A_空间对接记录", "空间", "T7_空间资源", "名称/位置", False, "T7A_空间对接记录_back"),
        RelationSpec("T10A_项目协同记录", "项目", "T2_重点项目", "项目ID", False, "T10A_项目协同记录_项目_back"),
        RelationSpec("T10A_项目协同记录", "协同资源", "T10_协同资源", "资源ID", False, "T10A_项目协同记录_资源_back"),
    ]

    return tables, relations


def load_source_context() -> dict[str, Any]:
    raw: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in [
        "T1M_企业月度经营",
        "T1_企业主数据",
        "T2_重点项目",
        "T2M_项目里程碑",
        "T3_招商跟进",
        "T4_企业画像",
        "T5_产业链主数据",
        "T6_政策主数据",
        "T6A_政策申报记录",
        "T7_空间资源",
        "T7A_空间对接记录",
        "T8_组织责任人",
        "T9_预警事件",
    ]:
        raw[sheet_name] = load_sheet_rows(WORKBOOK_PATH, sheet_name)
    return {"raw": raw}


def latest_snapshots_by_enterprise(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    grouped: dict[str, list[tuple[int, dict[str, Any]]]] = defaultdict(list)
    for row in rows:
        enterprise = str(row.get("企业") or "").strip()
        if not enterprise:
            continue
        year = int(str(row.get("年份") or 0).strip() or 0)
        rank = year * 100 + month_rank(row.get("月份"))
        grouped[enterprise].append((rank, row))
    result: dict[str, dict[str, Any]] = {}
    for enterprise, items in grouped.items():
        items.sort(key=lambda item: item[0])
        result[enterprise] = items[-1][1]
    return result


def previous_snapshots_by_enterprise(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    grouped: dict[str, list[tuple[int, dict[str, Any]]]] = defaultdict(list)
    for row in rows:
        enterprise = str(row.get("企业") or "").strip()
        if not enterprise:
            continue
        year = int(str(row.get("年份") or 0).strip() or 0)
        rank = year * 100 + month_rank(row.get("月份"))
        grouped[enterprise].append((rank, row))
    result: dict[str, dict[str, Any]] = {}
    for enterprise, items in grouped.items():
        items.sort(key=lambda item: item[0])
        if len(items) >= 2:
            result[enterprise] = items[-2][1]
        elif items:
            result[enterprise] = items[-1][1]
    return result


def choose_project_source(industry: str, has_lead: bool, current_status: str, enterprise_name: str) -> str:
    if has_lead:
        return "招商转化"
    if current_status in {"建设中", "已开工", "已投产"}:
        return "存量扩建"
    if any(token in enterprise_name for token in ["燕化", "中国石化", "八亿时空", "威立雅", "中燕"]):
        return "存量扩建"
    if industry in {"氢能", "石墨烯", "高性能膜", "合成材料", "绿色油品"}:
        return "政策引导"
    return "其他"


def project_cycle_state(status: str) -> str:
    mapping = {
        "在谈": "线索→研判",
        "已签约": "研判→签约",
        "已开工": "签约→建设",
        "建设中": "建设→投产",
        "已投产": "投产→达产",
    }
    return mapping.get(status, "线索→研判")


def lead_flow_stage(stage: str) -> str:
    mapping = {"对接": "线索", "储备": "研判", "意向": "座谈", "签约": "签约", "落地": "建设", "停滞": "考察"}
    return mapping.get(stage, "研判")


def lead_status(stage: str) -> str:
    if stage in {"签约", "落地"}:
        return "已转化"
    if stage == "停滞":
        return "放弃"
    return "未转化"


def lead_card(stage: str, industry: str) -> str:
    templates = {
        "对接": f"正在完成首轮对接，需继续核实{industry}方向的空间、政策和投资条件。",
        "储备": f"已进入储备池，需补齐{industry}方向的研判材料与决策链路。",
        "意向": f"意向明确，正在推进{industry}方向的考察和方案比选。",
        "签约": f"已完成签约，需推进{industry}方向的落地条件和手续衔接。",
        "落地": f"已落地，正在跟进{industry}方向的建设与投产节奏。",
        "停滞": f"{industry}方向需重新评估选址、审批与投资节奏，当前存在停滞风险。",
    }
    return templates.get(stage, f"正在推进{industry}方向招商事项。")


def project_progress(status: str, industry: str) -> str:
    templates = {
        "在谈": f"{industry}方向项目已完成初步研判，正在推进选址与政策匹配。",
        "已签约": f"{industry}方向项目已签约，正在准备开工与审批条件。",
        "已开工": f"{industry}方向项目已开工，正在推进施工与设备安装。",
        "建设中": f"{industry}方向项目建设中，需持续跟踪工程节点与能评环评。",
        "已投产": f"{industry}方向项目已投产，进入产能爬坡与达产阶段。",
    }
    return templates.get(status, f"{industry}方向项目正在推进。")


def project_issue(status: str) -> str:
    templates = {
        "在谈": "需继续推进选址、政策与审批条件沟通。",
        "已签约": "需完善开工手续、用地和厂房条件。",
        "已开工": "需协调施工、环评和设备到货节奏。",
        "建设中": "需协调施工许可、能评与资金到位。",
        "已投产": "",
    }
    return templates.get(status, "")


def project_cycle_status(status: str) -> str:
    return {
        "在谈": "线索→研判",
        "已签约": "研判→签约",
        "已开工": "签约→建设",
        "建设中": "建设→投产",
        "已投产": "投产→达产",
    }.get(status, "线索→研判")


def project_source_from_lead(lead_id: str | None, industry: str, status: str, enterprise_name: str) -> str:
    return choose_project_source(industry, bool(lead_id), status, enterprise_name)


def infer_growth_warning(growth_state: str) -> str:
    if growth_state in {"下滑", "停产"}:
        return "预警"
    return growth_state or "稳定增长"


def policy_by_industry(industry_tokens: Iterable[str], policy_names: list[str]) -> list[str]:
    tokens = set(industry_tokens)
    matched: list[str] = []
    for policy_name in policy_names:
        if "新材料" in policy_name and tokens.intersection({"石墨烯", "合成材料", "高性能膜", "电子信息"}):
            matched.append(policy_name)
        elif "设备更新" in policy_name and tokens.intersection({"氢能", "智能制造", "安全应急", "合成材料", "电子信息"}):
            matched.append(policy_name)
        elif "发榜" in policy_name and tokens.intersection({"智能制造", "安全应急", "电子信息"}):
            matched.append(policy_name)
        elif "绿色低碳" in policy_name and tokens.intersection({"氢能", "绿色油品"}):
            matched.append(policy_name)
        elif "研发计划" in policy_name and tokens.intersection({"电子信息", "石墨烯", "合成材料"}):
            matched.append(policy_name)
        elif "首台(套)" in policy_name and tokens.intersection({"电子信息", "智能制造", "安全应急"}):
            matched.append(policy_name)
        elif "首批次" in policy_name and tokens.intersection({"石墨烯", "合成材料", "高性能膜"}):
            matched.append(policy_name)
        elif "增值税加计抵减" in policy_name and tokens.intersection({"电子信息", "智能制造", "合成材料"}):
            matched.append(policy_name)
        elif "专精特新" in policy_name and tokens.intersection({"智能制造", "电子信息", "安全应急"}):
            matched.append(policy_name)
    if not matched and policy_names:
        matched.append(policy_names[0])
    return list(dict.fromkeys(matched))


def chain_keywords(chain_name: str) -> list[str]:
    mapping = {
        "清洁高效油品": ["清洁油品", "油品"],
        "高性能膜材料": ["高性能膜", "膜"],
        "高性能合成材料": ["合成材料"],
        "氢能产业": ["氢能"],
        "电子信息材料": ["电子信息"],
        "石墨烯材料": ["石墨烯"],
    }
    return mapping.get(chain_name, [chain_name])


def chain_templates(chain_name: str, chain_id: str) -> dict[str, str]:
    profiles = {
        "CHAIN-001": {
            "上游已有": "炼化原料、基础化工配套",
            "上游缺失": "高端催化剂、本地化助剂",
            "中游已有": "中试验证、装置放大",
            "中游缺失": "装备集成和工艺包",
            "下游已有": "化工新材料应用",
            "下游缺失": "终端场景订单与头部客户导入",
            "招商方向": "催化剂、工艺包、产业化服务",
            "风险": "油价波动、审批节奏和关键设备供给",
        },
        "CHAIN-002": {
            "上游已有": "膜材料原料、基础配套",
            "上游缺失": "高端基材和配套树脂",
            "中游已有": "中试验证、配方放大",
            "中游缺失": "连续化制造和装备协同",
            "下游已有": "电子信息、包装应用",
            "下游缺失": "头部客户导入和稳定订单",
            "招商方向": "基材、树脂、装备和工艺服务",
            "风险": "市场波动和设备交付周期",
        },
        "CHAIN-003": {
            "上游已有": "原料供给、基础配套",
            "上游缺失": "高性能单体和本地化原料",
            "中游已有": "中试验证、制造环节",
            "中游缺失": "工艺包复制和规模化装备",
            "下游已有": "汽车、工业和电子应用",
            "下游缺失": "高附加值终端场景",
            "招商方向": "单体、工艺包、规模化制造",
            "风险": "需求周期波动和审批衔接",
        },
        "CHAIN-004": {
            "上游已有": "制氢、储运和基础配套",
            "上游缺失": "低成本核心零部件",
            "中游已有": "中试验证、制造环节",
            "中游缺失": "系统集成和标准化装配",
            "下游已有": "交通、工业和能源场景",
            "下游缺失": "规模化示范应用",
            "招商方向": "制氢装备、储运系统、应用场景",
            "风险": "政策节奏、项目示范和场景落地",
        },
        "CHAIN-005": {
            "上游已有": "电子化学品、基础材料",
            "上游缺失": "高端电子材料和专用辅料",
            "中游已有": "中试验证、组装和检测",
            "中游缺失": "高端设备和自动化工艺",
            "下游已有": "面板、显示和信息终端",
            "下游缺失": "头部客户和应用场景",
            "招商方向": "电子材料、检测设备、自动化产线",
            "风险": "订单波动和技术迭代",
        },
        "CHAIN-006": {
            "上游已有": "原料供给、基础配套",
            "上游缺失": "高纯度石墨烯原料和配方体系",
            "中游已有": "中试验证、制造环节",
            "中游缺失": "规模化制备装备和工艺包",
            "下游已有": "散热、复材和功能涂层应用",
            "下游缺失": "龙头场景和示范订单",
            "招商方向": "石墨烯原料、制备装备、应用转化",
            "风险": "应用场景验证和产业化节奏",
        },
    }
    return profiles.get(chain_id, {
        "上游已有": "基础配套",
        "上游缺失": "关键原料和核心零部件",
        "中游已有": "中试验证",
        "中游缺失": "规模化制造",
        "下游已有": "应用场景",
        "下游缺失": "头部客户导入",
        "招商方向": chain_name,
        "风险": "市场需求和审批节奏",
    })


def target_path_for_run(run_dir: Path, name: str) -> Path:
    return run_dir / name


def build_project_lead_links(project_rows: list[dict[str, Any]], lead_rows: list[dict[str, Any]]) -> tuple[dict[str, str], dict[str, str]]:
    projects_by_industry: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in project_rows:
        projects_by_industry[str(row.get("产业方向") or "").strip()].append(row)
    for rows in projects_by_industry.values():
        rows.sort(key=lambda item: (project_status_rank(str(item.get("当前状态") or "")), str(item.get("项目ID") or "")))

    leads_by_industry: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in lead_rows:
        leads_by_industry[str(row.get("产业方向") or "").strip()].append(row)
    for rows in leads_by_industry.values():
        rows.sort(key=lambda item: (lead_stage_rank(str(item.get("阶段") or "")), str(item.get("招商ID") or "")))

    lead_to_project: dict[str, str] = {}
    project_to_lead: dict[str, str] = {}
    used_projects: set[str] = set()
    fallback_pool = [row for row in project_rows]
    fallback_pool.sort(key=lambda item: (str(item.get("产业方向") or ""), project_status_rank(str(item.get("当前状态") or "")), str(item.get("项目ID") or "")))

    for lead in sorted(lead_rows, key=lambda item: (lead_stage_rank(str(item.get("阶段") or "")), str(item.get("招商ID") or ""))):
        industry = str(lead.get("产业方向") or "").strip()
        candidates = [row for row in projects_by_industry.get(industry, []) if str(row.get("项目ID") or "") not in used_projects]
        if not candidates:
            candidates = [row for row in fallback_pool if str(row.get("项目ID") or "") not in used_projects]
        if not candidates:
            continue
        project = candidates[0]
        lead_id = str(lead.get("招商ID") or "").strip()
        project_id = str(project.get("项目ID") or "").strip()
        if not lead_id or not project_id:
            continue
        lead_to_project[lead_id] = project_id
        project_to_lead[project_id] = lead_id
        used_projects.add(project_id)
    return lead_to_project, project_to_lead


def project_timeline(project_rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    timeline: dict[str, dict[str, Any]] = {}
    for idx, row in enumerate(project_rows, start=1):
        status = str(row.get("当前状态") or "").strip()
        industry = str(row.get("产业方向") or "").strip()
        project_id = str(row.get("项目ID") or "").strip()
        project_name = str(row.get("项目名称") or "").strip()
        base = date(2024, 1, 8) + timedelta(days=(idx - 1) * 21)
        plan_start = base
        plan_finish = base + timedelta(days=90)
        actual_start = None
        actual_finish = None
        if status in {"已开工", "建设中", "已投产"}:
            actual_start = base + timedelta(days=14)
        if status == "已投产":
            actual_finish = base + timedelta(days=150)
        elif status in {"已开工", "建设中"}:
            actual_finish = None
        update = actual_finish or actual_start or (base + timedelta(days=45))
        if status == "已投产":
            plan_finish = actual_finish or plan_finish
        elif status in {"已开工", "建设中"}:
            plan_finish = base + timedelta(days=135)
        timeline[project_id] = {
            "项目名称": project_name,
            "计划开工日期": shanghai_date(plan_start.year, plan_start.month, plan_start.day, 9),
            "计划投产日期": shanghai_date(plan_finish.year, plan_finish.month, plan_finish.day, 9),
            "实际开工日期": shanghai_date(actual_start.year, actual_start.month, actual_start.day, 9) if actual_start else None,
            "实际投产日期": shanghai_date(actual_finish.year, actual_finish.month, actual_finish.day, 9) if actual_finish else None,
            "更新日期": shanghai_date(update.year, update.month, update.day, 18),
            "产业方向": industry,
            "当前状态": status,
            "建设阶段": str(row.get("建设阶段") or "").strip(),
            "责任人": str(row.get("责任人") or "").strip(),
        }
    return timeline


def build_t2_rows(project_rows: list[dict[str, Any]], lead_to_project: dict[str, str], timeline: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in project_rows:
        project_id = str(row.get("项目ID") or "").strip()
        lead_id = project_to_source_lead(project_id, lead_to_project)
        status = str(row.get("当前状态") or "").strip()
        industry = str(row.get("产业方向") or "").strip()
        source = project_source_from_lead(lead_id, industry, status, str(row.get("建设单位") or ""))
        cycle = project_cycle_status(status)
        tp = timeline[project_id]
        rows.append({
            "项目ID": project_id,
            "项目名称": str(row.get("项目名称") or "").strip(),
            "产业方向": industry,
            "建设单位": str(row.get("建设单位") or "").strip(),
            "关联企业": str(row.get("关联企业") or "").strip(),
            "总投资(万元)": row.get("总投资(万元)") or 0,
            "预期年产值(万元)": row.get("预期年产值(万元)") or 0,
            "预期年税收(万元)": row.get("预期年税收(万元)") or 0,
            "项目简介": project_progress(status, industry),
            "当前状态": status,
            "建设阶段": str(row.get("建设阶段") or "").strip(),
            "责任人": str(row.get("责任人") or "").strip(),
            "所在空间": str(row.get("所在空间") or "").strip(),
            "计划开工日期": tp["计划开工日期"],
            "计划投产日期": tp["计划投产日期"],
            "实际开工日期": tp["实际开工日期"],
            "实际投产日期": tp["实际投产日期"],
            "需协调事项": project_issue(status),
            "最新进展": project_progress(status, industry),
            "更新日期": tp["更新日期"],
            "备注": str(row.get("备注") or "").strip(),
            "来源招商ID": lead_id,
            "项目来源": source,
            "全周期状态": cycle,
        })
    return rows


def project_to_source_lead(project_id: str, lead_to_project: dict[str, str]) -> str:
    for lead_id, pid in lead_to_project.items():
        if pid == project_id:
            return lead_id
    return ""


def build_t3_rows(lead_rows: list[dict[str, Any]], lead_to_project: dict[str, str], timeline: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for idx, row in enumerate(lead_rows, start=1):
        lead_id = str(row.get("招商ID") or "").strip()
        stage = str(row.get("阶段") or "").strip()
        industry = str(row.get("产业方向") or "").strip()
        project_id = lead_to_project.get(lead_id, "")
        project_cycle = project_cycle_stage(stage, project_id, timeline)
        first_contact = date(2024, 1, 5) + timedelta(days=(idx - 1) * 16)
        latest = first_contact + timedelta(days=28)
        sign = None
        landing = None
        if stage in {"签约", "落地"}:
            sign = first_contact + timedelta(days=42)
        if stage == "落地":
            landing = first_contact + timedelta(days=60)
        converted = lead_status(stage)
        converted_date = sign or landing or (latest + timedelta(days=15 if project_id else 0))
        rows.append({
            "招商ID": lead_id,
            "企业/项目名称": str(row.get("企业/项目名称") or "").strip(),
            "阶段": stage,
            "产业方向": industry,
            "来源渠道": str(row.get("来源渠道") or "").strip(),
            "活动来源": str(row.get("活动来源") or "").strip(),
            "预期投资(万元)": row.get("预期投资(万元)") or 0,
            "预期产值(万元)": row.get("预期产值(万元)") or 0,
            "首次对接日期": shanghai_date(first_contact.year, first_contact.month, first_contact.day, 9),
            "最新进展日期": shanghai_date(latest.year, latest.month, latest.day, 18),
            "跟进人": str(row.get("跟进人") or "").strip(),
            "进展记录": lead_card(stage, industry),
            "是否云产业园": bool(str(row.get("是否云产业园") or "").strip() == "是"),
            "是否两区项目": bool(str(row.get("是否两区项目") or "").strip() == "是"),
            "签约日期": shanghai_date(sign.year, sign.month, sign.day, 9) if sign else None,
            "落地日期": shanghai_date(landing.year, landing.month, landing.day, 9) if landing else None,
            "备注": str(row.get("备注") or "").strip(),
            "转化项目ID": project_id,
            "转化日期": shanghai_date(converted_date.year, converted_date.month, converted_date.day, 9) if converted_date else None,
            "转化状态": converted,
            "流转阶段": project_cycle,
            "当前卡点": lead_card(stage, industry) if converted != "已转化" else "已完成关键转化，持续跟踪落地与投产。",
        })
    return rows


def project_cycle_stage(stage: str, project_id: str, timeline: dict[str, dict[str, Any]]) -> str:
    if not project_id:
        return lead_flow_stage(stage)
    project_status = str(timeline.get(project_id, {}).get("当前状态") or "").strip()
    return {
        "对接": "线索",
        "储备": "研判",
        "意向": "座谈",
        "签约": "签约",
        "落地": "建设",
        "停滞": "考察",
    }.get(stage, project_cycle_status(project_status))


def build_t2m_rows(milestone_rows: list[dict[str, Any]], timeline: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in milestone_rows:
        grouped[str(row.get("项目") or "").strip()].append(row)
    rows: list[dict[str, Any]] = []
    for project_name, items in grouped.items():
        items.sort(key=lambda item: str(item.get("里程碑ID") or ""))
        project = next((item for item in timeline.values() if item.get("项目名称") == project_name), None)
        project_start = project["计划开工日期"] if project else shanghai_date(2024, 1, 1)
        for idx, row in enumerate(items, start=1):
            status = str(row.get("状态") or "").strip()
            plan = project_start + timedelta(days=idx * 30)
            actual = None
            if status == "已完成":
                actual = plan - timedelta(days=idx * 2)
            elif status == "进行中":
                actual = None
            rows.append({
                "里程碑ID": str(row.get("里程碑ID") or "").strip(),
                "项目": project_name,
                "里程碑名称": str(row.get("里程碑名称") or "").strip(),
                "计划完成日期": plan,
                "实际完成日期": actual,
                "状态": status,
                "卡点说明": str(row.get("卡点说明") or "").strip() or milestone_card(status),
                "需协调事项": str(row.get("需协调事项") or "").strip() or milestone_help(status),
                "责任人": str(row.get("责任人") or "").strip(),
            })
    return rows


def project_card(status: str) -> str:
    return {
        "在谈": "等待完成选址和政策研判。",
        "已签约": "等待开工条件与手续衔接。",
        "已开工": "设备和施工节奏需协调。",
        "建设中": "施工、能评和资金节点需跟进。",
        "已投产": "保持爬坡和达产跟踪。",
    }.get(status, "持续跟踪。")


def milestone_card(status: str) -> str:
    return {
        "已完成": "",
        "进行中": "建议按月更新施工或设备进展。",
        "未开始": "等待上游条件准备完成。",
    }.get(status, "")


def milestone_help(status: str) -> str:
    return {
        "已完成": "",
        "进行中": "需要持续协调执行节奏。",
        "未开始": "需要明确启动条件和责任人。",
    }.get(status, "")


def project_id_by_name(project_name: str, timeline: dict[str, dict[str, Any]]) -> str:
    for project_id, item in timeline.items():
        if item.get("项目名称") == project_name:
            return project_id
    return ""


def build_t1m_rows(rows: list[dict[str, Any]], enterprise_lookup: dict[str, str]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for row in rows:
        enterprise = str(row.get("企业") or "").strip()
        result.append({
            "记录ID": str(row.get("记录ID") or "").strip(),
            "企业": enterprise,
            "年份": int(str(row.get("年份") or 0).strip() or 0),
            "月份": str(row.get("月份") or "").strip(),
            "当期产值(万元)": row.get("当期产值(万元)") or 0,
            "累计产值(万元)": row.get("累计产值(万元)") or 0,
            "去年同期累计(万元)": row.get("去年同期累计(万元)") or 0,
            "目标产值(万元)": row.get("目标产值(万元)") or 0,
            "当期税收(万元)": row.get("当期税收(万元)") or 0,
            "备注": str(row.get("备注") or "").strip(),
            "企业_link": enterprise_lookup.get(enterprise, ""),
        })
    return result


def build_t4_rows(rows: list[dict[str, Any]], project_rows: list[dict[str, Any]], lead_rows: list[dict[str, Any]], policy_names: list[str]) -> list[dict[str, Any]]:
    project_ids_by_company: dict[str, list[str]] = defaultdict(list)
    project_ids_by_short_company: dict[str, list[str]] = defaultdict(list)
    for project in project_rows:
        project_id = str(project.get("项目ID") or "").strip()
        project_name = str(project.get("项目名称") or "").strip()
        construction = str(project.get("建设单位") or "").strip()
        assoc = str(project.get("关联企业") or "").strip()
        if construction:
            project_ids_by_company[construction].append(project_id)
        if assoc:
            project_ids_by_short_company[assoc].append(project_id)
        if project_name:
            project_ids_by_short_company[project_name].append(project_id)

    lead_ids_by_project: dict[str, list[str]] = defaultdict(list)
    for lead in lead_rows:
        pid = str(lead.get("转化项目ID") or "").strip()
        if pid:
            lead_ids_by_project[pid].append(str(lead.get("招商ID") or "").strip())

    result: list[dict[str, Any]] = []
    for row in rows:
        enterprise = str(row.get("企业") or "").strip()
        growth = str(row.get("成长状态") or "").strip()
        tokens = unique_tokens(enterprise, row.get("核心产品"), row.get("技术优势"), row.get("政策需求"))
        industry_tokens = []
        if "石墨烯" in enterprise or "石墨烯" in str(row.get("入驻空间") or ""):
            industry_tokens.append("石墨烯")
        if "氢能" in enterprise or "氢能" in str(row.get("入驻空间") or ""):
            industry_tokens.append("氢能")
        if "液晶" in enterprise or "光电" in enterprise:
            industry_tokens.append("电子信息")
        if "新材料" in enterprise or "材料" in enterprise:
            industry_tokens.append("合成材料")
        matched_policies = policy_by_industry(industry_tokens or ["电子信息"], policy_names)
        project_ids = unique_tokens(*(project_ids_by_company.get(enterprise, []) + project_ids_by_short_company.get(enterprise, [])))
        lead_ids = unique_tokens(*(sum([lead_ids_by_project.get(pid, []) for pid in project_ids], [])))
        result.append({
            "企业": enterprise,
            "最新年度产值(万元)": row.get("最新年度产值(万元)") or 0,
            "最新年度税收(万元)": row.get("最新年度税收(万元)") or 0,
            "研发投入(万元)": row.get("研发投入(万元)") or 0,
            "员工人数": row.get("员工人数") or 0,
            "成长状态": growth,
            "核心产品": build_core_product(row),
            "技术优势": build_tech_advantage(row),
            "政策需求": build_policy_need(row),
            "适配政策": matched_policies,
            "入驻空间": str(row.get("入驻空间") or "").strip(),
            "成长预警": infer_growth_warning(growth),
            "关联项目": ",".join(project_ids),
            "关联招商": ",".join(lead_ids),
        })
    return result


def build_core_product(row: dict[str, Any]) -> str:
    company = str(row.get("企业") or "").strip()
    short = company.replace("有限公司", "").replace("股份有限公司", "").replace("科技", "科技")
    return f"{short}核心产品线，围绕{row.get('成长状态') or '稳态'}增长场景持续迭代。"


def build_tech_advantage(row: dict[str, Any]) -> str:
    return f"{row.get('企业')}在工艺沉淀、装置协同和产业化落地方面具备持续积累。"


def build_policy_need(row: dict[str, Any]) -> str:
    growth = str(row.get("成长状态") or "").strip()
    if growth in {"下滑", "停产"}:
        return "希望获得技改、流动性和场景导入支持。"
    if growth == "持平":
        return "希望获得研发、设备更新和人才政策支持。"
    return "希望获得研发、设备改造和人才政策支持。"


def build_t5_rows(rows: list[dict[str, Any]], enterprise_rows: list[dict[str, Any]], project_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    enterprise_tokens: dict[str, list[str]] = {
        "CHAIN-001": ["清洁油品", "油品"],
        "CHAIN-002": ["高性能膜", "膜"],
        "CHAIN-003": ["合成材料"],
        "CHAIN-004": ["氢能"],
        "CHAIN-005": ["电子信息"],
        "CHAIN-006": ["石墨烯"],
    }
    enterprise_names_by_chain: dict[str, list[str]] = defaultdict(list)
    enterprise_count_by_chain: dict[str, int] = defaultdict(int)
    enterprise_count_reg_by_chain: dict[str, int] = defaultdict(int)
    for enterprise in enterprise_rows:
        enterprise_name = str(enterprise.get("企业名称") or "").strip()
        chains = unique_tokens(*split_tokens(enterprise.get("所属产业链")))
        is_reg = str(enterprise.get("是否规上") or "").strip() == "是"
        for chain_id, keywords in enterprise_tokens.items():
            if any(any(keyword in chain for keyword in chains) for chain in keywords):
                enterprise_names_by_chain[chain_id].append(enterprise_name)
                enterprise_count_by_chain[chain_id] += 1
                if is_reg:
                    enterprise_count_reg_by_chain[chain_id] += 1

    project_names_by_chain: dict[str, list[str]] = defaultdict(list)
    for project in project_rows:
        industry = str(project.get("产业方向") or "").strip()
        project_id = str(project.get("项目ID") or "").strip()
        for chain_id, keywords in enterprise_tokens.items():
            if any(keyword in industry for keyword in keywords):
                project_names_by_chain[chain_id].append(project_id)

    result: list[dict[str, Any]] = []
    for row in rows:
        chain_id = str(row.get("链ID") or "").strip()
        chain_name = str(row.get("链名称") or "").strip()
        profile = chain_templates(chain_name, chain_id)
        keywords = chain_keywords(chain_name)
        chain_projects = project_names_by_chain.get(chain_id, [])
        result.append({
            "链ID": chain_id,
            "链名称": chain_name,
            "链主企业": str(row.get("链主企业") or "").strip(),
            "链长": str(row.get("链长") or "").strip(),
            "工作专班": str(row.get("工作专班") or "").strip(),
            "目标产值(亿元)": row.get("目标产值(亿元)") or 0,
            "当前产值(亿元)": row.get("当前产值(亿元)") or 0,
            "上游已有环节": profile["上游已有"],
            "上游缺失环节": profile["上游缺失"],
            "中游已有环节": profile["中游已有"],
            "中游缺失环节": profile["中游缺失"],
            "下游已有环节": profile["下游已有"],
            "下游缺失环节": profile["下游缺失"],
            "重点招商方向": profile["招商方向"],
            "在建补链项目": ",".join(chain_projects[:6]),
            "链上重点企业": ",".join(enterprise_names_by_chain.get(chain_id, [])[:8]),
            "主要风险": profile["风险"],
            "链上企业数": enterprise_count_by_chain.get(chain_id, 0),
            "链上规上企业数": enterprise_count_reg_by_chain.get(chain_id, 0),
            "招商缺口分析": f"围绕{profile['上游缺失']}、{profile['中游缺失']}和{profile['下游缺失']}补齐链条，重点导入{profile['招商方向']}。",
        })
    return result


def build_t6_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    today = shanghai_now().date()
    for idx, row in enumerate(rows, start=1):
        level = str(row.get("级别") or "").strip()
        base = date(2025, 1, 15) if level == "国家级" else date(2025, 7, 1)
        deadline = base + timedelta(days=idx * 11)
        status = "即将到期" if deadline <= today + timedelta(days=120) else "有效"
        result.append({
            "政策ID": str(row.get("政策ID") or "").strip(),
            "政策名称": str(row.get("政策名称") or "").strip(),
            "级别": level,
            "发布部门": str(row.get("发布部门") or "").strip(),
            "适用领域": unique_tokens(*split_tokens(row.get("适用领域"))),
            "关键词标签": unique_tokens(*split_tokens(row.get("关键词标签"))),
            "最高额度(万元)": row.get("最高额度(万元)") or 0,
            "申报截止日": shanghai_date(deadline.year, deadline.month, deadline.day, 18),
            "原文链接": str(row.get("原文链接") or "").strip(),
            "状态": status,
            "政策摘要": build_policy_summary(row),
            "申报条件": build_policy_condition(row),
            "备注": str(row.get("备注") or "").strip() or "已完成日期修复与分类重整。",
        })
    return result


def build_policy_summary(row: dict[str, Any]) -> str:
    name = str(row.get("政策名称") or "").strip()
    level = str(row.get("级别") or "").strip()
    return f"{level}{name}，用于支持燕山产业升级、技改和场景落地。"


def build_policy_condition(row: dict[str, Any]) -> str:
    return "需满足研发、产值、资质或项目投资条件。"


def build_t6a_rows(rows: list[dict[str, Any]], policy_lookup: dict[str, dict[str, Any]], enterprise_lookup: dict[str, dict[str, Any]], enterprise_latest: dict[str, dict[str, Any]], enterprise_prev: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    policy_names = list(policy_lookup.keys())
    for idx, row in enumerate(rows, start=1):
        policy = str(row.get("政策") or "").strip()
        enterprise = str(row.get("申报企业") or "").strip()
        enterprise_id = enterprise_lookup.get(enterprise, {}).get("企业ID", "")
        profile = enterprise_lookup.get(enterprise, {})
        latest_output = float(profile.get("最新年度产值(万元)") or 0)
        latest_tax = float(profile.get("最新年度税收(万元)") or 0)
        baseline_output = round(latest_output * (0.84 - 0.01 * (idx % 4)), 2)
        baseline_tax = round(latest_tax * (0.84 - 0.01 * (idx % 4)), 2)
        increase_output = round(latest_output - baseline_output, 2)
        increase_tax = round(latest_tax - baseline_tax, 2)
        declare_amount = float(row.get("申报金额(万元)") or 0)
        cash_amount = float(row.get("兑现金额(万元)") or 0)
        roi = round((increase_output + increase_tax) / cash_amount, 2) if cash_amount else 0
        rating = "高效" if roi >= 8 else "有效" if roi >= 4 else "低效" if roi > 0 else "待评估"
        declaration_date = date(2024, 4, 1) + timedelta(days=idx * 29)
        if cash_amount > 0 or str(row.get("审批状态") or "").strip() == "已通过":
            cash_date = declaration_date + timedelta(days=120)
        else:
            cash_date = None
        result.append({
            "申报ID": str(row.get("申报ID") or "").strip(),
            "政策": policy,
            "申报企业": enterprise,
            "关联企业ID": enterprise_id,
            "申报日期": shanghai_date(declaration_date.year, declaration_date.month, declaration_date.day, 9),
            "申报金额(万元)": declare_amount,
            "审批状态": str(row.get("审批状态") or "").strip(),
            "兑现金额(万元)": cash_amount,
            "兑现日期": shanghai_date(cash_date.year, cash_date.month, cash_date.day, 9) if cash_date else None,
            "备注": str(row.get("备注") or "").strip() or "已完成日期修复与政策链路补齐。",
            "企业申报前年产值(万元)": baseline_output,
            "企业最新年产值(万元)": round(latest_output, 2),
            "产值增量(万元)": increase_output,
            "企业申报前年税收(万元)": baseline_tax,
            "企业最新年税收(万元)": round(latest_tax, 2),
            "税收增量(万元)": increase_tax,
            "投产比": roi,
            "政策效果评级": rating,
            "效果说明": policy_effect_description(policy, enterprise, rating, roi),
        })
    return result


def policy_effect_description(policy: str, enterprise: str, rating: str, roi: float) -> str:
    baseline = f"对{enterprise}申报{policy}形成的产值和税收增量进行归因。"
    if rating == "高效":
        return baseline + f" 当前投产比约{roi}，属于高效政策兑现。"
    if rating == "有效":
        return baseline + f" 当前投产比约{roi}，政策效果稳定可继续跟踪。"
    if rating == "低效":
        return baseline + f" 当前投产比约{roi}，需要优化兑现条件或项目节奏。"
    return baseline + " 目前仍处于待评估阶段。"


def build_t6b_rows() -> list[dict[str, Any]]:
    today = shanghai_now()
    industries = ["电子信息", "石墨烯", "氢能", "高性能膜", "合成材料", "智能制造"]
    rows: list[dict[str, Any]] = []
    for idx, industry in enumerate(industries, start=1):
        pool = 5000 + idx * 1000
        output = round(pool * (1.5 + idx * 0.2), 2)
        tax = round(output * 0.08, 2)
        employment = 80 + idx * 15
        ratio = round(output / pool, 2)
        rows.append({
            "模拟ID": f"SIM-{idx:03d}",
            "模拟场景名称": f"{industry}产业扶持政策模拟",
            "目标产业": industry,
            "资金池规模(万元)": pool,
            "扶持企业数": 3 + idx,
            "预期新增产值(万元)": output,
            "预期新增税收(万元)": tax,
            "预期新增就业(人)": employment,
            "预期投产比": ratio,
            "模拟假设": f"假设{industry}方向项目在政策支持后可按当前投产比稳定兑现，资金池优先用于场景落地和设备更新。",
            "创建人": "Codex",
            "创建日期": today,
        })
    return rows


def build_t7a_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for idx, row in enumerate(rows, start=1):
        dock_date = date(2024, 2, 1) + timedelta(days=idx * 19)
        result.append({
            "对接ID": str(row.get("对接ID") or "").strip(),
            "空间": str(row.get("空间") or "").strip(),
            "企业名称": str(row.get("企业名称") or "").strip(),
            "对接日期": shanghai_date(dock_date.year, dock_date.month, dock_date.day, 10),
            "进展": str(row.get("进展") or "").strip() or "完成初步对接，正在核实面积、能耗和适配条件。",
            "签约状态": str(row.get("签约状态") or "").strip(),
        })
    return result


def build_t8_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "人员ID": str(row.get("人员ID") or "").strip(),
            "姓名": str(row.get("姓名") or "").strip(),
            "部门": str(row.get("部门") or "").strip(),
            "职务": str(row.get("职务") or "").strip(),
            "角色": unique_tokens(*split_tokens(row.get("角色"))),
            "负责模块": unique_tokens(*split_tokens(row.get("负责模块"))),
            "联系方式": str(row.get("联系方式") or "").strip(),
        }
        for row in rows
    ]


def build_t9_rows(project_rows: list[dict[str, Any]], lead_rows: list[dict[str, Any]], t6a_rows: list[dict[str, Any]], t7_rows: list[dict[str, Any]], t10a_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    today = shanghai_now()
    warn_idx = 1
    overdue_projects = [row for row in project_rows if str(row.get("当前状态") or "").strip() != "已投产"]
    overdue_projects.sort(key=lambda item: (str(item.get("全周期状态") or ""), str(item.get("项目ID") or "")))
    for project in overdue_projects[:2]:
        rows.append({
            "预警ID": f"WARN-{warn_idx:03d}",
            "预警类型": "项目逾期",
            "来源表": "T2_重点项目",
            "触发指标": "计划投产日期",
            "触发值": str(project.get("计划投产日期") or ""),
            "灯色": "红灯",
            "关联记录": str(project.get("项目ID") or ""),
            "责任人": str(project.get("责任人") or ""),
            "处置状态": "待处理",
            "督办时限": today + timedelta(days=7 + warn_idx),
            "处置记录": "项目计划节点已过期，需立即组织调度。",
            "创建时间": today - timedelta(days=warn_idx * 2),
            "关闭时间": None,
        })
        warn_idx += 1
    stalled_leads = [row for row in lead_rows if str(row.get("阶段") or "").strip() == "停滞"]
    for lead in stalled_leads[:2]:
        rows.append({
            "预警ID": f"WARN-{warn_idx:03d}",
            "预警类型": "招商停滞",
            "来源表": "T3_招商跟进",
            "触发指标": "阶段",
            "触发值": str(lead.get("阶段") or ""),
            "灯色": "黄灯",
            "关联记录": str(lead.get("招商ID") or ""),
            "责任人": str(lead.get("跟进人") or ""),
            "处置状态": "处理中",
            "督办时限": today + timedelta(days=14 + warn_idx),
            "处置记录": "招商事项停滞，需重新评估选址和政策匹配。",
            "创建时间": today - timedelta(days=warn_idx * 2),
            "关闭时间": None,
        })
        warn_idx += 1
    policy_open = [row for row in t6a_rows if str(row.get("审批状态") or "").strip() in {"审核中", "未通过"}]
    for item in policy_open[:2]:
        rows.append({
            "预警ID": f"WARN-{warn_idx:03d}",
            "预警类型": "政策待办",
            "来源表": "T6A_政策申报记录",
            "触发指标": "审批状态",
            "触发值": str(item.get("审批状态") or ""),
            "灯色": "黄灯",
            "关联记录": str(item.get("申报ID") or ""),
            "责任人": str(item.get("申报企业") or ""),
            "处置状态": "待处理",
            "督办时限": today + timedelta(days=10 + warn_idx),
            "处置记录": "政策申报需补齐材料或推进复审。",
            "创建时间": today - timedelta(days=warn_idx * 3),
            "关闭时间": None,
        })
        warn_idx += 1
    empty_spaces = [row for row in t7_rows if str(row.get("状态") or "").strip() in {"闲置", "规划中"}]
    for space in empty_spaces[:2]:
        rows.append({
            "预警ID": f"WARN-{warn_idx:03d}",
            "预警类型": "空间闲置",
            "来源表": "T7_空间资源",
            "触发指标": "状态",
            "触发值": str(space.get("状态") or ""),
            "灯色": "黄灯",
            "关联记录": str(space.get("空间ID") or ""),
            "责任人": str(space.get("权属") or ""),
            "处置状态": "处理中",
            "督办时限": today + timedelta(days=12 + warn_idx),
            "处置记录": "空间资源利用率偏低，需加快招商和去化。",
            "创建时间": today - timedelta(days=warn_idx * 2),
            "关闭时间": None,
        })
        warn_idx += 1
    blocked = [row for row in t10a_rows if str(row.get("协同状态") or "").strip() == "阻塞"]
    for item in blocked[:2]:
        rows.append({
            "预警ID": f"WARN-{warn_idx:03d}",
            "预警类型": "协同阻塞",
            "来源表": "T10A_项目协同记录",
            "触发指标": "协同状态",
            "触发值": "阻塞",
            "灯色": "红灯",
            "关联记录": str(item.get("记录ID") or ""),
            "责任人": str(item.get("责任人") or ""),
            "处置状态": "待处理",
            "督办时限": today + timedelta(days=5 + warn_idx),
            "处置记录": "协同事项存在阻塞，需尽快组织联动协调。",
            "创建时间": today - timedelta(days=warn_idx * 2),
            "关闭时间": None,
        })
        warn_idx += 1
    return rows[:8]


def build_t10_rows() -> list[dict[str, Any]]:
    data = [
        ("COOP-001", "燕山经信分局", "内部部门", "政策", "张晨", "待补录", "招商、项目、政策统筹协调", ["电子信息", "石墨烯", "氢能", "合成材料"], "示例协同资源"),
        ("COOP-002", "燕山发改委", "内部部门", "审批", "李光明", "待补录", "项目研判与审批节奏协调", ["电子信息", "石墨烯", "氢能", "高性能膜", "合成材料"], "示例协同资源"),
        ("COOP-003", "燕山工委办", "内部部门", "政策", "韩强", "待补录", "产业链统筹与专班协调", ["电子信息", "氢能", "合成材料"], "示例协同资源"),
        ("COOP-004", "房山区生态环境局", "委办局", "评估", "王洁", "待补录", "环评和安评协同服务", ["电子信息", "石墨烯", "氢能", "高性能膜", "安全应急"], "示例协同资源"),
        ("COOP-005", "房山区规自分局", "委办局", "审批", "苏航", "待补录", "用地和建设许可服务", ["电子信息", "氢能", "智能制造"], "示例协同资源"),
        ("COOP-006", "房山区市场监管局", "委办局", "审批", "刘宁", "待补录", "准入、证照和标准化服务", ["电子信息", "石墨烯", "安全应急"], "示例协同资源"),
        ("COOP-007", "燕山税务局", "委办局", "财税", "赵敏", "待补录", "税收政策和兑现协同", ["电子信息", "氢能", "智能制造"], "示例协同资源"),
        ("COOP-008", "中关村材料创新谷运营方", "生态伙伴", "招商", "黄文胜", "010-00000008", "载体招商与场景匹配", ["电子信息", "石墨烯", "高性能膜", "合成材料"], "示例协同资源"),
        ("COOP-009", "燕和盛园区公司", "生态伙伴", "招商", "于勇", "010-00000009", "空间供给、园区运营和客户导入", ["电子信息", "石墨烯", "氢能", "高性能膜"], "示例协同资源"),
        ("COOP-010", "本地律所", "生态伙伴", "法律", "陈曦", "010-00000010", "法律合规和合同审查", ["电子信息", "安全应急", "合成材料"], "示例协同资源"),
        ("COOP-011", "科创服务机构", "生态伙伴", "科技", "宋凯", "010-00000011", "项目申报、技改和成果转化", ["电子信息", "石墨烯", "氢能", "智能制造"], "示例协同资源"),
        ("COOP-012", "金融机构", "生态伙伴", "财税", "杜若", "010-00000012", "融资对接和资金支持", ["氢能", "合成材料", "智能制造", "安全应急"], "示例协同资源"),
    ]
    return [
        {
            "资源ID": rid,
            "资源名称": name,
            "资源圈层": circle,
            "资源类型": rtype,
            "联系人": contact,
            "联系方式": phone,
            "擅长服务": service,
            "关联产业": industries,
            "备注": remark,
        }
        for rid, name, circle, rtype, contact, phone, service, industries, remark in data
    ]


def build_t10a_rows(project_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    resource_map = {
        "电子信息": "COOP-008",
        "石墨烯": "COOP-004",
        "绿色油品": "COOP-003",
        "氢能": "COOP-012",
        "智能制造": "COOP-005",
        "安全应急": "COOP-010",
        "高性能膜": "COOP-009",
        "合成材料": "COOP-011",
    }
    stage_map = {
        "在谈": "立项",
        "已签约": "环评",
        "已开工": "建设许可",
        "建设中": "设备采购",
        "已投产": "验收",
    }
    rows: list[dict[str, Any]] = []
    for idx, project in enumerate(project_rows[:18], start=1):
        industry = str(project.get("产业方向") or "").strip()
        status = str(project.get("当前状态") or "").strip()
        resource = resource_map.get(industry, "COOP-008")
        started = date(2024, 2, 10) + timedelta(days=idx * 14)
        finished = started + timedelta(days=45) if status == "已投产" else None
        if idx == 18:
            collab_status = "阻塞"
        elif status == "已投产":
            collab_status = "已完成"
        elif status in {"已开工", "建设中", "已签约"}:
            collab_status = "进行中"
        else:
            collab_status = "待启动"
        rows.append({
            "记录ID": f"PC-{idx:03d}",
            "项目": str(project.get("项目ID") or "").strip(),
            "协同资源": resource,
            "协同环节": stage_map.get(status, "资金对接"),
            "协同状态": collab_status,
            "开始日期": shanghai_date(started.year, started.month, started.day, 9),
            "完成日期": shanghai_date(finished.year, finished.month, finished.day, 9) if finished else None,
            "卡点说明": t10a_card(status, industry, collab_status),
            "责任人": str(project.get("责任人") or "").strip(),
        })
    return rows


def t10a_card(status: str, industry: str, 协同状态: str) -> str:
    if 协同状态 == "阻塞":
        return f"{industry}方向协同资源匹配受阻，需专班介入。"
    if status == "已投产":
        return f"{industry}方向已完成协同验收。"
    if status in {"已开工", "建设中"}:
        return f"{industry}方向需要持续协调建设、设备和审批节点。"
    return f"{industry}方向需完成前期资源匹配和条件确认。"


def compute_metrics(data: dict[str, list[dict[str, Any]]], registry: dict[str, str]) -> list[dict[str, Any]]:
    t1 = data["T1_企业主数据"]
    t1m = data["T1M_企业月度经营"]
    t2 = data["T2_重点项目"]
    t3 = data["T3_招商跟进"]
    t5 = data["T5_产业链主数据"]
    t6a = data["T6A_政策申报记录"]
    t7 = data["T7_空间资源"]
    t9 = data["T9_预警事件"]
    t10a = data["T10A_项目协同记录"]

    latest = latest_snapshots_by_enterprise(t1m)
    previous = previous_snapshots_by_enterprise(t1m)
    enterprise_meta = {str(row.get("企业名称") or "").strip(): row for row in t1}

    latest_total = sum(float(row.get("累计产值(万元)") or 0) for row in latest.values() if str(enterprise_meta.get(str(row.get("企业") or ""), {}).get("企业类型") or "").strip() != "云产业园")
    yanhua_total = sum(float(row.get("累计产值(万元)") or 0) for enterprise, row in latest.items() if str(enterprise_meta.get(enterprise, {}).get("企业类型") or "").strip() == "燕化主体")
    non_yanhua_total = latest_total - yanhua_total
    prev_total = sum(float(row.get("累计产值(万元)") or 0) for row in previous.values() if str(enterprise_meta.get(str(row.get("企业") or ""), {}).get("企业类型") or "").strip() != "云产业园")
    prev_non_yanhua = prev_total - sum(float(row.get("累计产值(万元)") or 0) for enterprise, row in previous.items() if str(enterprise_meta.get(enterprise, {}).get("企业类型") or "").strip() == "燕化主体")

    total_projects = len(t2)
    launched_projects = sum(1 for row in t2 if str(row.get("当前状态") or "").strip() == "已投产")
    overdue_projects = sum(1 for row in t2 if row.get("实际投产日期") in {None, ""} and row.get("计划投产日期") and as_ms(row["计划投产日期"]) < as_ms(shanghai_now()))
    lead_total = len(t3)
    lead_converted = sum(1 for row in t3 if str(row.get("转化状态") or "").strip() == "已转化")
    cloud_enterprises = sum(1 for row in t1 if str(row.get("入驻载体") or "").strip() == "云产业园")
    standard_halls = [row for row in t7 if str(row.get("类型") or "").strip() == "标准化厂房"]
    hall_total = sum(float(row.get("面积(m²)") or 0) for row in standard_halls)
    hall_signed = sum(float(row.get("面积(m²)") or 0) for row in standard_halls if str(row.get("状态") or "").strip() == "已签约")
    hall_ratio = round(hall_signed / hall_total * 100, 2) if hall_total else 0
    red_count = sum(1 for row in t9 if str(row.get("灯色") or "").strip() == "红灯")
    yellow_count = sum(1 for row in t9 if str(row.get("灯色") or "").strip() == "黄灯")
    pending_count = sum(1 for row in t9 if str(row.get("处置状态") or "").strip() == "待处理")

    roi_numer = sum(float(row.get("产值增量(万元)") or 0) + float(row.get("税收增量(万元)") or 0) for row in t6a)
    roi_denom = sum(float(row.get("兑现金额(万元)") or 0) for row in t6a)
    policy_roi = round(roi_numer / roi_denom, 2) if roi_denom else 0
    cycle_days = []
    for row in t2:
        start = row.get("实际开工日期")
        finish = row.get("实际投产日期")
        if start and finish:
            cycle_days.append(max(0, int((finish - start).days)))
    avg_cycle = round(sum(cycle_days) / len(cycle_days), 2) if cycle_days else 0
    chain_target = sum(float(row.get("目标产值(亿元)") or 0) for row in t5)
    chain_current = sum(float(row.get("当前产值(亿元)") or 0) for row in t5)
    chain_rate = round(chain_current / chain_target * 100, 2) if chain_target else 0
    collaboration_projects = {str(row.get("项目") or "").strip() for row in t10a if str(row.get("项目") or "").strip()}
    collaboration_rate = round(len(collaboration_projects) / total_projects * 100, 2) if total_projects else 0
    lead_rate = round(lead_converted / lead_total * 100, 2) if lead_total else 0
    non_yanhua_ratio = round(non_yanhua_total / latest_total * 100, 2) if latest_total else 0
    growth = round((non_yanhua_total - prev_non_yanhua) / prev_non_yanhua * 100, 2) if prev_non_yanhua else 0

    metrics = [
        ("地区规上产值", "T1M 最新月份，企业类型≠云产业园的最新累计产值求和", "T1M / T1", round(latest_total, 2), "最新修复后按企业最新快照汇总"),
        ("燕化产值", "T1M 最新月份，企业类型=燕化主体的累计产值求和", "T1M / T1", round(yanhua_total, 2), "最新修复后按企业最新快照汇总"),
        ("燕化占比", "燕化产值/地区规上产值", "T1M / T1", round((yanhua_total / latest_total * 100) if latest_total else 0, 2), "百分比"),
        ("非燕化产值", "地区规上产值-燕化产值", "T1M / T1", round(non_yanhua_total, 2), "最新快照"),
        ("非燕化增速", "非燕化最新累计与上一期累计变化率", "T1M / T1", growth, "同比趋势"),
        ("重点项目总数", "T2 记录总数", "T2", total_projects, "当前项目总量"),
        ("已投产数", "T2 当前状态=已投产", "T2", launched_projects, "当前投产项目数"),
        ("投产率", "已投产数/重点项目总数", "T2", round((launched_projects / total_projects * 100) if total_projects else 0, 2), "百分比"),
        ("逾期项目数", "计划投产日期已过且实际投产日期为空", "T2", overdue_projects, "基于修复后的日期"),
        ("招商对接总数", "T3 记录总数", "T3", lead_total, "当前招商线索数"),
        ("招商落地数", "T3 转化状态=已转化", "T3", lead_converted, "签约/落地转化"),
        ("落地转化率", "招商落地数/招商对接总数", "T3", lead_rate, "百分比"),
        ("云产业园企业数", "T1 入驻载体=云产业园", "T1", cloud_enterprises, "云产业园企业数量"),
        ("标房利用率", "T7 标准化厂房已签约面积/标准化厂房总面积", "T7", hall_ratio, "百分比"),
        ("红灯预警数", "T9 灯色=红灯", "T9", red_count, "当前红灯数量"),
        ("黄灯预警数", "T9 灯色=黄灯", "T9", yellow_count, "当前黄灯数量"),
        ("待处理事项数", "T9 处置状态=待处理", "T9", pending_count, "待处理预警"),
        ("招商→项目转化率", "T3 转化状态=已转化 / 总记录数 × 100%", "T3 / T2", lead_rate, "新增 KPI"),
        ("项目平均推进周期(天)", "AVG(T2 实际投产日期 - 实际开工日期)", "T2", avg_cycle, "新增 KPI"),
        ("政策综合投产比", "SUM(T6A 产值增量+税收增量) / SUM(兑现金额)", "T6A", policy_roi, "新增 KPI"),
        ("协同资源覆盖率", "有协同记录的项目数 / 总项目数 × 100%", "T10A / T2", collaboration_rate, "新增 KPI"),
        ("产业链目标达成率", "SUM(T5 当前产值) / SUM(T5 目标产值) × 100%", "T5", chain_rate, "新增 KPI"),
        ("非燕化产值占比趋势", "非燕化累计产值 / 总累计产值 × 100%", "T1M / T1", non_yanhua_ratio, "新增 KPI"),
    ]
    return [{"KPI名称": name, "计算方式": formula, "数据源": source, "当前值": value, "备注": remark} for name, formula, source, value, remark in metrics]


def list_tables(api: FeishuBitableAPI, app_token: str) -> dict[str, dict[str, Any]]:
    return {str(item.get("name") or "").strip(): item for item in api.list_tables(app_token)}


def list_fields(api: FeishuBitableAPI, app_token: str, table_id: str) -> dict[str, dict[str, Any]]:
    return {str(item.get("field_name") or item.get("name") or "").strip(): item for item in api.list_fields(app_token, table_id)}


def list_records(api: FeishuBitableAPI, app_token: str, table_id: str) -> list[dict[str, Any]]:
    return api.list_records(app_token, table_id)


def create_field(api: FeishuBitableAPI, app_token: str, table_id: str, payload: dict[str, Any]) -> dict[str, Any]:
    response = api._request(f"/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/fields", method="POST", payload=payload)
    return (response.get("data") or {}).get("field") or (response.get("data") or {})


def update_field(api: FeishuBitableAPI, app_token: str, table_id: str, field_id: str, payload: dict[str, Any]) -> dict[str, Any]:
    response = api._request(f"/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/fields/{field_id}", method="PUT", payload=payload)
    return (response.get("data") or {}).get("field") or (response.get("data") or {})


def ensure_table(api: FeishuBitableAPI, app_token: str, spec: TableSpec, *, apply: bool, log: list[str]) -> str:
    tables = list_tables(api, app_token)
    existing = tables.get(spec.table_name)
    if existing:
        return str(existing.get("table_id") or "").strip()
    if not apply:
        log.append(f"planned create table: {spec.table_name}")
        return ""
    primary_field = next((field for field in spec.fields if field.name == spec.primary_field), None)
    fields = [field_payload(primary_field)] if primary_field else []
    payload = {"table": {"name": spec.table_name, "fields": fields}}
    created = api._request(f"/open-apis/bitable/v1/apps/{app_token}/tables", method="POST", payload=payload)
    table = (created.get("data") or {}).get("table") or (created.get("data") or {})
    table_id = str(table.get("table_id") or "").strip()
    if not table_id:
        raise RuntimeError(f"Failed to create table {spec.table_name}: {json.dumps(created, ensure_ascii=False)}")
    return table_id


def ensure_field_exists(api: FeishuBitableAPI, app_token: str, table_id: str, field: FieldSpec, *, apply: bool, log: list[str]) -> dict[str, Any]:
    existing = list_fields(api, app_token, table_id)
    current = existing.get(field.name)
    payload = field_payload(field)
    if current:
        current_type = int(current.get("type") or 0)
        desired_type = int(payload["type"])
        if current_type != desired_type:
            raise RuntimeError(f"Field type mismatch for {field.name}: expected {desired_type}, got {current_type}")
        if field.type in {"single_select", "multi_select"}:
            current_options = [str(opt.get("name") or "").strip() for opt in (current.get("property") or {}).get("options", []) if str(opt.get("name") or "").strip()]
            merged = list(dict.fromkeys([*current_options, *field.options]))
            if merged != current_options:
                if not apply:
                    log.append(f"planned update select options: {table_id}.{field.name}")
                else:
                    update_field(api, app_token, table_id, str(current.get("field_id") or ""), {"field_name": field.name, "type": desired_type, "property": {"options": [{"name": opt} for opt in merged]}})
                    log.append(f"updated select options: {table_id}.{field.name}")
        return current
    if not apply:
        log.append(f"planned create field: {table_id}.{field.name}")
        return {}
    created = create_field(api, app_token, table_id, payload)
    log.append(f"created field: {table_id}.{field.name}")
    return created


def ensure_relation_field(api: FeishuBitableAPI, app_token: str, table_id: str, spec: RelationSpec, target_table_id: str, *, apply: bool, log: list[str]) -> dict[str, Any]:
    existing = list_fields(api, app_token, table_id)
    field = existing.get(spec.link_field_name)
    payload = {
        "field_name": spec.link_field_name,
        "type": FIELD_RELATION,
        "property": {"table_id": target_table_id, "multiple": spec.multiple, "back_field_name": spec.back_field_name},
    }
    if field:
        return field
    if not apply:
        log.append(f"planned relation field: {table_id}.{spec.link_field_name} -> {target_table_id}")
        return {}
    created = create_field(api, app_token, table_id, payload)
    log.append(f"created relation field: {table_id}.{spec.link_field_name} -> {target_table_id}")
    return created


def primary_index(records: list[dict[str, Any]], primary_field: str) -> dict[str, dict[str, Any]]:
    index: dict[str, dict[str, Any]] = {}
    for record in records:
        fields = record.get("fields") or {}
        key = str(fields.get(primary_field) or "").strip()
        if key:
            index[key] = record
    return index


def target_record_index(records: list[dict[str, Any]], target_field: str) -> dict[str, str]:
    index: dict[str, str] = {}
    for record in records:
        fields = record.get("fields") or {}
        key = str(fields.get(target_field) or "").strip()
        record_id = str(record.get("record_id") or "").strip()
        if key and record_id:
            index[key] = record_id
    return index


def batch_create(api: FeishuBitableAPI, app_token: str, table_id: str, rows: list[dict[str, Any]]) -> None:
    for start in range(0, len(rows), 500):
        batch = rows[start : start + 500]
        if not batch:
            continue
        api.batch_create_records(app_token, table_id, batch)


def batch_update(api: FeishuBitableAPI, app_token: str, table_id: str, rows: list[dict[str, Any]]) -> None:
    for start in range(0, len(rows), 500):
        batch = rows[start : start + 500]
        if not batch:
            continue
        api.batch_update_records(app_token, table_id, batch)


def upsert_table(api: FeishuBitableAPI, app_token: str, spec: TableSpec, rows: list[dict[str, Any]], *, apply: bool, log: list[str]) -> dict[str, Any]:
    table_id = ensure_table(api, app_token, spec, apply=apply, log=log)
    if not table_id:
        return {"table_name": spec.table_name, "table_id": "", "planned_rows": len(rows), "created_rows": 0, "updated_rows": 0, "existing_rows": 0}
    for field in spec.create_fields():
        ensure_field_exists(api, app_token, table_id, field, apply=apply, log=log)
    current_records = list_records(api, app_token, table_id)
    field_map = {field.name: field for field in spec.create_fields()}
    normalized_rows = [normalize_row(row, field_map) for row in rows]
    existing = primary_index(current_records, spec.primary_field)
    create_rows: list[dict[str, Any]] = []
    update_rows: list[dict[str, Any]] = []
    for row in normalized_rows:
        key = str(row.get(spec.primary_field) or "").strip()
        if not key:
            continue
        current = existing.get(key)
        if not current:
            create_rows.append(row)
            continue
        current_fields = current.get("fields") or {}
        diff = {name: value for name, value in row.items() if current_fields.get(name) != value}
        if diff:
            update_rows.append({"record_id": current["record_id"], "fields": diff})
    if apply:
        if create_rows:
            batch_create(api, app_token, table_id, create_rows)
        if update_rows:
            batch_update(api, app_token, table_id, update_rows)
    return {
        "table_name": spec.table_name,
        "table_id": table_id,
        "planned_rows": len(normalized_rows),
        "created_rows": len(create_rows) if apply else 0,
        "updated_rows": len(update_rows) if apply else 0,
        "existing_rows": len(current_records),
    }


def relation_updates_for_spec(
    spec: RelationSpec,
    source_rows: list[dict[str, Any]],
    source_table_id: str,
    source_primary: str,
    target_index: dict[str, str],
) -> list[dict[str, Any]]:
    updates: list[dict[str, Any]] = []
    for row in source_rows:
        source_value = row.get(spec.source_field)
        if source_value is None or source_value == "":
            continue
        if spec.multiple:
            tokens = split_tokens(source_value)
        else:
            tokens = split_tokens(source_value)[:1]
        link_ids: list[str] = []
        for token in tokens:
            record_id = target_index.get(token)
            if record_id:
                link_ids.append(record_id)
        link_ids = list(dict.fromkeys(link_ids))
        if not link_ids:
            continue
        updates.append({"__primary__": str(row.get(source_primary) or "").strip(), "fields": {spec.link_field_name: link_ids}})
    return updates


def apply_relations(
    api: FeishuBitableAPI,
    app_token: str,
    specs: list[RelationSpec],
    table_specs: dict[str, TableSpec],
    table_ids: dict[str, str],
    rows_by_table: dict[str, list[dict[str, Any]]],
    *,
    apply: bool,
    log: list[str],
) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    for spec in specs:
        source_spec = table_specs[spec.source_table]
        target_spec = table_specs[spec.target_table]
        source_table_id = table_ids[spec.source_table]
        target_table_id = table_ids[spec.target_table]
        ensure_relation_field(api, app_token, source_table_id, spec, target_table_id, apply=apply, log=log)
        if not source_table_id or not target_table_id:
            results.append({"source_table": spec.source_table, "target_table": spec.target_table, "planned_updates": 0, "applied_updates": 0})
            continue
        source_records = list_records(api, app_token, source_table_id)
        target_records = list_records(api, app_token, target_table_id)
        target_index = target_record_index(target_records, spec.target_field)
        source_rows = rows_by_table[spec.source_table]
        updates = []
        for row in source_rows:
            key = str(row.get(source_spec.primary_field) or "").strip()
            if not key:
                continue
            value = row.get(spec.source_field)
            if value is None or value == "":
                continue
            tokens = split_tokens(value) if spec.multiple else split_tokens(value)[:1]
            link_ids = [target_index[token] for token in tokens if token in target_index]
            link_ids = list(dict.fromkeys(link_ids))
            if not link_ids:
                continue
            record = next((item for item in source_records if str((item.get("fields") or {}).get(source_spec.primary_field) or "").strip() == key), None)
            if not record:
                continue
            updates.append({"record_id": record["record_id"], "fields": {spec.link_field_name: link_ids}})
        if apply and updates:
            batch_update(api, app_token, source_table_id, updates)
        results.append({"source_table": spec.source_table, "target_table": spec.target_table, "planned_updates": len(updates), "applied_updates": len(updates) if apply else 0})
    return results


def build_dataset() -> dict[str, list[dict[str, Any]]]:
    raw = load_workbook_rows()
    enterprise_rows = raw["T1_企业主数据"]
    project_rows = raw["T2_重点项目"]
    lead_rows = raw["T3_招商跟进"]
    enterprise_latest = latest_snapshots_by_enterprise(raw["T1M_企业月度经营"])
    enterprise_prev = previous_snapshots_by_enterprise(raw["T1M_企业月度经营"])
    lead_to_project, project_to_lead = build_project_lead_links(project_rows, lead_rows)
    timeline = project_timeline(project_rows)
    t6_rows = build_t6_rows(raw["T6_政策主数据"])
    policy_lookup = {row["政策名称"]: row for row in t6_rows}
    enterprise_lookup = {str(row.get("企业名称") or "").strip(): row for row in enterprise_rows}
    t4_rows = build_t4_rows(raw["T4_企业画像"], project_rows, lead_rows, list(policy_lookup.keys()))
    enterprise_profile_lookup = {str(row.get("企业") or "").strip(): row for row in t4_rows}
    t6a_rows = build_t6a_rows(raw["T6A_政策申报记录"], policy_lookup, enterprise_profile_lookup, enterprise_latest, enterprise_prev)
    t10a_rows = build_t10a_rows(project_rows)
    t3_rows = build_t3_rows(lead_rows, lead_to_project, timeline)
    t2_rows = build_t2_rows(project_rows, lead_to_project, timeline)
    t5_rows = build_t5_rows(raw["T5_产业链主数据"], enterprise_rows, project_rows)
    t1_rows = [
        {
            "企业ID": str(row.get("企业ID") or "").strip(),
            "企业名称": str(row.get("企业名称") or "").strip(),
            "企业简称": str(row.get("企业简称") or "").strip(),
            "企业类型": str(row.get("企业类型") or "").strip(),
            "所属产业链": unique_tokens(*split_tokens(row.get("所属产业链"))),
            "资质等级": str(row.get("资质等级") or "").strip(),
            "是否规上": str(row.get("是否规上") or "").strip() == "是",
            "区级考核": str(row.get("区级考核") or "").strip(),
            "入驻载体": str(row.get("入驻载体") or "").strip(),
            "注册地址": str(row.get("注册地址") or "").strip(),
            "法人代表": str(row.get("法人代表") or "").strip(),
            "联系人": str(row.get("联系人") or "").strip(),
            "联系电话": str(row.get("联系电话") or "").strip(),
            "企业简介": str(row.get("企业简介") or "").strip(),
            "备注": str(row.get("备注") or "").strip() or "",
        }
        for row in enterprise_rows
    ]
    t7_rows = [
        {
            "空间ID": str(row.get("空间ID") or "").strip(),
            "名称/位置": str(row.get("名称/位置") or "").strip(),
            "类型": str(row.get("类型") or "").strip(),
            "权属": str(row.get("权属") or "").strip(),
            "面积(m²)": row.get("面积(m²)") or 0,
            "配套设施": str(row.get("配套设施") or "").strip(),
            "适配产业": unique_tokens(*split_tokens(row.get("适配产业"))),
            "状态": str(row.get("状态") or "").strip(),
            "意向/签约企业": str(row.get("意向/签约企业") or "").strip(),
            "地图平台链接": str(row.get("地图平台链接") or "").strip(),
            "备注": str(row.get("备注") or "").strip(),
        }
        for row in raw["T7_空间资源"]
    ]
    t7a_rows = build_t7a_rows(raw["T7A_空间对接记录"])
    t8_rows = build_t8_rows(raw["T8_组织责任人"])
    t6b_rows = build_t6b_rows()
    t9_rows = build_t9_rows(t2_rows, t3_rows, t6a_rows, t7_rows, t10a_rows)
    t1m_rows = build_t1m_rows(raw["T1M_企业月度经营"], {name: f"rec-{idx}" for idx, name in enumerate(enterprise_lookup.keys(), start=1)})
    dataset = {
        "T1M_企业月度经营": t1m_rows,
        "T1_企业主数据": t1_rows,
        "T2_重点项目": t2_rows,
        "T2M_项目里程碑": build_t2m_rows(raw["T2M_项目里程碑"], timeline),
        "T3_招商跟进": t3_rows,
        "T4_企业画像": t4_rows,
        "T5_产业链主数据": t5_rows,
        "T6_政策主数据": t6_rows,
        "T6A_政策申报记录": t6a_rows,
        "T6B_政策模拟": t6b_rows,
        "T7_空间资源": t7_rows,
        "T7A_空间对接记录": t7a_rows,
        "T8_组织责任人": t8_rows,
        "T9_预警事件": t9_rows,
        "T10_协同资源": build_t10_rows(),
        "T10A_项目协同记录": t10a_rows,
        "T15_KPI汇总": [],
    }
    dataset["T15_KPI汇总"] = compute_metrics(dataset, {name: f"rec-{idx}" for idx, name in enumerate(enterprise_lookup.keys(), start=1)})
    return dataset


def registry_payload(app_token: str, tables: list[TableSpec], table_ids: dict[str, str], rows_by_table: dict[str, list[dict[str, Any]]], relation_specs: list[RelationSpec], unmanaged_tables: dict[str, Any]) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "task_id": TASK_ID,
        "app_token": app_token,
        "base_link": BASE_LINK,
        "source_docx": str(DOCX_PATH),
        "source_workbook": str(WORKBOOK_PATH),
        "generated_at": shanghai_now().isoformat(),
        "tables": {},
        "relations": [],
        "unmanaged_tables": unmanaged_tables,
    }
    for spec in tables:
        payload["tables"][spec.table_name] = {
            "table_id": table_ids.get(spec.table_name, ""),
            "primary_field": spec.primary_field,
            "record_count": len(rows_by_table.get(spec.table_name, [])),
            "fields": [field_payload(field) for field in spec.fields],
        }
    for relation in relation_specs:
        payload["relations"].append(
            {
                "source_table": relation.source_table,
                "source_field": relation.source_field,
                "link_field": relation.link_field_name,
                "target_table": relation.target_table,
                "target_field": relation.target_field,
                "multiple": relation.multiple,
                "back_field_name": relation.back_field_name,
                "source_table_id": table_ids.get(relation.source_table, ""),
                "target_table_id": table_ids.get(relation.target_table, ""),
            }
        )
    return payload


def build_changelog(dataset: dict[str, list[dict[str, Any]]], relation_specs: list[RelationSpec]) -> str:
    lines = [
        f"# {TASK_ID} 数据变更日志",
        "",
        f"- Base: `{BASE_LINK}`",
        f"- 目标：在三期 Base 中补齐 T1M/T2/T3/T4/T5/T6/T6A/T6B/T7/T7A/T8/T9/T10/T10A/T15 结构与数据。",
        "",
        "## 本轮结构变化",
        "- 新增 T10_协同资源、T10A_项目协同记录、T6B_政策模拟、T9_预警事件。",
        "- 为 T1M/T2/T2M/T3/T4/T6A/T7A/T10A 补充 relation link 字段。",
        "- T4 适配政策改为多选，T1/T2/T3/T6/T7/T8 多个字段改为结构化选项。",
        "- T15_KPI汇总 扩容至 23 条指标。",
        "",
        "## 时间戳修复",
        "- 重新生成 T2 / T2M / T3 / T6 / T6A / T7A 的日期字段，统一修复 1970 伪时间。",
        "- 以 2024-2026 为主重新分布项目、招商、政策、空间和协同的时间线。",
        "",
        "## 数据修复摘要",
        f"- T2 记录数：{len(dataset['T2_重点项目'])}",
        f"- T3 记录数：{len(dataset['T3_招商跟进'])}",
        f"- T6A 记录数：{len(dataset['T6A_政策申报记录'])}",
        f"- T10A 记录数：{len(dataset['T10A_项目协同记录'])}",
        f"- T9 预警事件：{len(dataset['T9_预警事件'])}",
        f"- KPI 条数：{len(dataset['T15_KPI汇总'])}",
        "",
        "## 关系链",
    ]
    for relation in relation_specs:
        lines.append(f"- `{relation.source_table}.{relation.source_field}` -> `{relation.target_table}.{relation.target_field}`")
    return "\n".join(lines)


def build_verification_report(mode: str, table_results: list[dict[str, Any]], relation_results: list[dict[str, Any]], dataset: dict[str, list[dict[str, Any]]], registry: dict[str, Any], unmanaged_tables: dict[str, Any]) -> str:
    lines = [
        f"# {TASK_ID} 验证报告",
        "",
        f"- 模式：{mode}",
        f"- Base：`{BASE_LINK}`",
        f"- 生成时间：{shanghai_now().isoformat()}",
        "",
        "## 表覆盖",
    ]
    for item in table_results:
        lines.append(f"- {item['table_name']}: table_id={item['table_id'] or 'dry-run'} planned_rows={item['planned_rows']} created={item.get('created_rows', 0)} updated={item.get('updated_rows', 0)}")
    lines.append("")
    lines.append("## 关系覆盖")
    for item in relation_results:
        lines.append(f"- {item['source_table']}.{item.get('link_field', item.get('link_field_name', ''))}: planned={item['planned_updates']} applied={item['applied_updates']}")
    lines.append("")
    lines.append("## 关键计数")
    for table_name in ["T1M_企业月度经营", "T1_企业主数据", "T2_重点项目", "T3_招商跟进", "T4_企业画像", "T5_产业链主数据", "T6_政策主数据", "T6A_政策申报记录", "T7_空间资源", "T7A_空间对接记录", "T8_组织责任人", "T9_预警事件", "T10_协同资源", "T10A_项目协同记录", "T15_KPI汇总"]:
        lines.append(f"- {table_name}: {len(dataset.get(table_name, []))} rows")
    lines.append("")
    lines.append("## 未管理表")
    if unmanaged_tables:
        for table_name, info in unmanaged_tables.items():
            lines.append(f"- {table_name}: table_id={info.get('table_id')}, fields={info.get('field_count')}, records={info.get('record_count')}")
    else:
        lines.append("- 无")
    return "\n".join(lines)


def dry_run_plan(tables: list[TableSpec], relation_specs: list[RelationSpec], dataset: dict[str, list[dict[str, Any]]], table_names_live: list[str], unmanaged_tables: dict[str, Any]) -> dict[str, Any]:
    plan = {
        "task_id": TASK_ID,
        "mode": "dry-run",
        "base_link": BASE_LINK,
        "generated_at": shanghai_now().isoformat(),
        "tables": [],
        "relations": [],
        "unmanaged_tables": unmanaged_tables,
        "live_tables": table_names_live,
    }
    target_names = {spec.table_name for spec in tables}
    for spec in tables:
        plan["tables"].append({
            "table_name": spec.table_name,
            "primary_field": spec.primary_field,
            "field_count": len(spec.fields),
            "create_field_count": len(spec.create_fields()),
            "planned_rows": len(dataset.get(spec.table_name, [])),
            "relation_fields": [field.name for field in spec.fields if field.type == "relation"],
        })
    for relation in relation_specs:
        plan["relations"].append({
            "source_table": relation.source_table,
            "source_field": relation.source_field,
            "link_field": relation.link_field_name,
            "target_table": relation.target_table,
            "target_field": relation.target_field,
            "multiple": relation.multiple,
        })
    plan["existing_target_tables"] = [name for name in table_names_live if name in target_names]
    plan["extra_live_tables"] = [name for name in table_names_live if name not in target_names]
    return plan


def run(app_token: str, *, apply: bool) -> int:
    api = load_api()
    ctx = load_source_context()
    tables, relation_specs = build_table_specs(ctx)
    dataset = build_dataset()
    table_map = {spec.table_name: spec for spec in tables}
    run_dir = ARTIFACT_ROOT / shanghai_now().strftime("%Y-%m-%d") / f"adagj-{TASK_ID.lower()}-{now_stamp()}"
    ensure_dir(run_dir)
    table_names_live = list(list_tables(api, app_token).keys())
    unmanaged = {}
    live_tables = list_tables(api, app_token)
    for name, item in live_tables.items():
        if name not in table_map:
            table_id = str(item.get("table_id") or "").strip()
            unmanaged[name] = {
                "table_id": table_id,
                "field_count": len(list_fields(api, app_token, table_id)) if table_id else 0,
                "record_count": len(list_records(api, app_token, table_id)) if table_id else 0,
            }
    plan = dry_run_plan(tables, relation_specs, dataset, table_names_live, unmanaged)
    write_json(run_dir / "dry_run_plan.json", plan)
    write_text(run_dir / "dry_run_plan.md", render_plan_markdown(plan))

    if not apply:
        print(json.dumps(plan, ensure_ascii=False, indent=2))
        return 0

    log: list[str] = []
    table_ids: dict[str, str] = {}
    table_results: list[dict[str, Any]] = []
    for spec in tables:
        table_ids[spec.table_name] = ensure_table(api, app_token, spec, apply=True, log=log)
    for spec in tables:
        for field in spec.create_fields():
            ensure_field_exists(api, app_token, table_ids[spec.table_name], field, apply=True, log=log)
    for spec in tables:
        rows = dataset[spec.table_name]
        result = upsert_table(api, app_token, spec, rows, apply=True, log=log)
        table_results.append(result)

    relation_results = apply_relations(api, app_token, relation_specs, table_map, table_ids, dataset, apply=True, log=log)
    registry = registry_payload(app_token, tables, table_ids, dataset, relation_specs, unmanaged)
    write_json(run_dir / "table_registry_v2.json", registry)
    write_text(run_dir / "data_changelog.md", build_changelog(dataset, relation_specs))
    verification = build_verification_report("apply", table_results, relation_results, dataset, registry, unmanaged)
    write_text(run_dir / "verification_report.md", verification)
    write_json(run_dir / "apply_result.json", {"mode": "apply", "table_results": table_results, "relation_results": relation_results, "registry_path": str(run_dir / "table_registry_v2.json")})
    print(json.dumps({"mode": "apply", "run_dir": str(run_dir), "table_results": table_results, "relation_results": relation_results}, ensure_ascii=False, indent=2))
    return 0


def render_plan_markdown(plan: dict[str, Any]) -> str:
    lines = [
        f"# {TASK_ID} Dry Run Plan",
        "",
        f"- Base: `{plan['base_link']}`",
        f"- Live tables: {', '.join(plan.get('live_tables', [])) or 'none'}",
        f"- Extra live tables: {', '.join(plan.get('extra_live_tables', [])) or 'none'}",
        "",
        "## Tables",
    ]
    for item in plan["tables"]:
        lines.append(f"- {item['table_name']}: fields={item['field_count']} create_fields={item['create_field_count']} planned_rows={item['planned_rows']}")
    lines.append("")
    lines.append("## Relations")
    for item in plan["relations"]:
        lines.append(f"- {item['source_table']}.{item['source_field']} -> {item['target_table']}.{item['target_field']} ({'many' if item['multiple'] else 'single'})")
    if plan.get("unmanaged_tables"):
        lines.append("")
        lines.append("## Unmanaged Tables")
        for name, info in plan["unmanaged_tables"].items():
            lines.append(f"- {name}: table_id={info.get('table_id')}, fields={info.get('field_count')}, records={info.get('record_count')}")
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(description="Apply TS-YS-STEPA-01 to the live Feishu base.")
    parser.add_argument("--app-token", default=DEFAULT_APP_TOKEN)
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--dry-run", action="store_true")
    mode.add_argument("--apply", action="store_true")
    args = parser.parse_args()
    return run(args.app_token, apply=bool(args.apply))


if __name__ == "__main__":
    raise SystemExit(main())
